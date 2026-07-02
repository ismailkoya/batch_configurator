"""Batch Teltonika Device Configurator (FMBxxx family).

Single-file Tkinter app for batch-configuring Teltonika devices.

User picks (once):
  - [x] Firmware Update     <browse>  <fixed filename>
  - [x] Config Upload       <browse>  <fixed filename>
  - Password: ( ) Set Keyword <new>
              ( ) Remove Keyword <current>
              (•) No changes

Flow per device:
  1. Detect (USB VID/PID filter + probe :cfg_info:?)
  2. 5-second cable-wiggle grace
  3. Authenticate if secured (using "current keyword" from form)
  4. Firmware update (.xim → TLTF unwrap → MMM minify → YMODEM-1K)
  5. Config upload (.cfg → gunzip → parse → batched :cfg_setparam → :cfg_save)
  6. Keyword Set / Remove
  7. Append row to grid + return to "waiting for next device"
  8. Reconnect: try last-known port first for 5s, fall back to full scan

Dependencies (pip install):
  - pyserial
  - cryptography     (for AES-256-CBC during .xim decrypt)

The .xim TLTF unwrap and MMM minify are PURE Python — no native DLL
needed. They mirror the WebCrypto / DataView port in our configurator
(index.html, "ximParse" + "_ximMinifyMmm"), themselves a port of
Teltonika's F# Reader + xim-minifier-x64.dll.

Wire-level fluency was lifted from configurator/index.html — boot
handshake, batched cfg_setparam, security commands all match official.
"""
from __future__ import annotations

import gzip
import hashlib
import os
import queue
import re
import struct
import sys
import threading
import time
import tkinter as tk
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Callable, Optional

try:
    import serial
    from serial.tools import list_ports
except ImportError:
    sys.exit("ERROR: pyserial not installed. Run:  pip install pyserial")

try:
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
except ImportError:
    sys.exit("ERROR: cryptography not installed. Run:  pip install cryptography")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run:  pip install openpyxl")


def _app_dir() -> Path:
    """Directory the app should treat as its "home" for side-by-side files
    (log file, defaults, etc.).

    When running from a PyInstaller `--onefile` build, `__file__` points at
    the temporary extraction folder (`%TEMP%\\_MEIxxxx\\`) which is deleted
    at process exit — writing the log there means it vanishes with the app.
    In that case we use the directory of the .exe itself (from
    `sys.executable`) so the log lands next to what the user actually
    double-clicked. In normal Python-source runs we keep the old behaviour
    of using the script's own directory.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

# openpyxl powers the XLSX export. We import it lazily-tolerant: dev runs
# without it still launch fine; the Export XLSX button will surface a clear
# install hint if it's missing. The packaged exe always has it because
# build.bat installs it before PyInstaller.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

# Teltonika USB device filters (VID:PID), from fmb_data.json. Each FMB
# device enumerates as TWO COM ports (one debug, one config) when fully
# initialised, so we expect to see PAIRS of these on the bus.
USB_DEVICE_FILTERS = [
    ("0E8D", "0023"),   # MediaTek CDC (FMB modem)
    ("0483", "5740"),   # STMicro CDC (FMB)
    ("1FC9", "0094"),   # NXP LPC CDC
    ("0483", "5702"),   # STMicro composite (MI_00)
    ("1D12", "0200"),   # Spec/Industrial CDC
]

# Model variant map ("slot2 raw" → user-facing).
# Full hardware-variant → display-model map lifted from the configurator's
# Teltonika.hardwareVersions table (71 entries covering every shipping FMB,
# FMC, FMM, FMP, FMT, FMU, FMB9x, FMC8x, FMM8x and MTB device). Built from
# globo360_teltonika.js so it stays consistent with the configurator we
# ship alongside this tool.
MODEL_VARIANT_MAP = {
    "FMB0:1":   "FMB010",
    "FMB0:2":   "FMB001",
    "FMB0:3":   "FMB001 Buzzer",
    "FMB0:4":   "FMB003",
    "FMB0:5":   "FMB002",
    "FMB0:6":   "FMB020",
    "FMB1:1":   "FMB110 DualSIM",
    "FMB1:2":   "FMB120, FMB122 DualSIM",
    "FMB1:3":   "FMB125 DualSIM",
    "FMB1:4":   "FMB125 no BAT DualSIM",
    "FMB1:5":   "FMB130",
    "FMB1:6":   "FMB140",
    "FMB1:7":   "FMB150",
    "FMB1:21":  "FMB110 BLE",
    "FMB1:22":  "FMB120, FMB122 BLE",
    "FMB1:23":  "FMB125",
    "FMB1:24":  "FMB125 no BAT",
    "FMB1:130": "FMB225",
    "FMB1:132": "FMB230",
    "FMB1:133": "FMB240",
    "FMB1:134": "FMB250",
    "FMB1:161": "FMB225 DualSIM",
    "FMB2:1":   "FMB202",
    "FMB2:2":   "FMB204",
    "FMB2:3":   "FMB206",
    "FMB9:7":   "FMB910",
    "FMB9:21":  "FMB900",
    "FMB9:22":  "FMB920",
    "FMC0:1":   "FMC001",
    "FMC0:3":   "FMC003",
    "FMC0:4":   "FMC00A",
    "FMC1:3":   "FMC125",
    "FMC1:4":   "FMC130",
    "FMC1:5":   "FMC13A",
    "FMC1:6":   "FMC150 2 DOUT",
    "FMC1:7":   "FMC150 3 DOUT",
    "FMC1:130": "FMC225",
    "FMC1:131": "FMC230",
    "FMC1:133": "FMC250 2 DOUT",
    "FMC1:134": "FMC250 3 DOUT",
    "FMC2:1":   "FMC234",
    "FMC8:1":   "FMC800",
    "FMC8:2":   "FMC880",
    "FMC9:1":   "FMC920",
    "FMM0:1":   "FMM001 BG96",
    "FMM0:2":   "FMM001 Buzzer",
    "FMM0:21":  "FMM001 BG95M3",
    "FMM0:22":  "FMM003 BG95M3",
    "FMM0:25":  "FMM003 BG95M3 Buzzer",
    "FMM0:31":  "FMM001 BG95M1",
    "FMM0:34":  "FMM00A BG95M1",
    "FMM0:41":  "FMM003 BG95M6",
    "FMM0:51":  "FMM003 BG96 Buzzer",
    "FMM1:3":   "FMM125 BG96",
    "FMM1:4":   "FMM130 BG96",
    "FMM1:5":   "FMM150",
    "FMM1:23":  "FMM125 BG95M3",
    "FMM1:24":  "FMM130 BG95M3",
    "FMM1:25":  "FMM13A",
    "FMM1:132": "FMM250",
    "FMM1:151": "FMM230 BG95M3",
    "FMM8:1":   "FMM800",
    "FMM8:2":   "FMM880",
    "FMM8:3":   "FMM80A",
    "FMM9:1":   "FMM920",
    "FMP1:1":   "FMP100",
    "FMT1:1":   "FMT100",
    "FMU1:3":   "FMU125",
    "FMU1:4":   "FMU130",
    "FMU1:5":   "FMU126",
    "MTB1:2":   "MTB100",
}

BAUD               = 115200
PORT_POLL_MS       = 300         # COM-port scan cadence
PRE_START_DELAY_S  = 5           # cable-wiggle grace after detect
LAST_PORT_WAIT_S   = 5           # retry the previous COM port for this long
PROBE_TIMEOUT_S    = 2.0
SECSTAT_TIMEOUT_S  = 3.0
SETPARAM_TIMEOUT_S = 15.0
CFG_SAVE_TIMEOUT_S = 10.0
SETPARAM_BATCH_B   = 1024

# ─────────────────────────────────────────────────────────────────────────────
# .xim → .bin unwrap (pure Python, ports configurator JS exactly)
# ─────────────────────────────────────────────────────────────────────────────

# Base64 of the 32-byte AES password — same constant as the JS.
_XIM_PASSWORD = bytes.fromhex(
    # "hX9Xe0ttKg0SOCPRbAjdAcPy6VD96NRYo9GQ00MrEfU=" in hex
    "857f577b4b6d2a0d123823d16c08dd01c3f2e950fde8d458a3d190d3432b11f5"
)


def _xim_aes_cbc_decrypt_iso10126(key: bytes, iv: bytes, ct: bytes) -> bytes:
    """AES-256-CBC decrypt with ISO 10126 padding (random tail).

    Standard cryptography lib only supports PKCS7. Same trick as the JS:
    encrypt a synthetic block that XORs to PKCS7-pad (0x10×16), append it,
    decrypt with PKCS7, then strip our junk + the original ISO10126 tail.
    """
    if len(ct) % 16:
        raise ValueError(".xim ciphertext length is not a multiple of 16")
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv))
    decryptor = cipher.decryptor()
    plain = decryptor.update(ct) + decryptor.finalize()
    # ISO10126 last byte = pad count (1..16); strip that many bytes.
    pad_n = plain[-1]
    if not 1 <= pad_n <= 16:
        raise ValueError(f"Bad ISO10126 padding: pad_n={pad_n}")
    return plain[:-pad_n]


def _xim_parse_tltf(xim: bytes) -> dict:
    """Parse TLTF container header: { salt, iv, ct, sig }."""
    if len(xim) < 16:
        raise ValueError("TLTF file too short")
    if xim[:4] != b"TLTF":
        raise ValueError("Not a TLTF file")
    if struct.unpack_from("<i", xim, 4)[0] != 1:
        raise ValueError("Unsupported TLTF version")
    off = 16
    def read(label):
        nonlocal off
        (n,) = struct.unpack_from("<i", xim, off); off += 4
        if n < 0 or off + n > len(xim):
            raise ValueError(f"Truncated TLTF {label}")
        b = xim[off:off + n]; off += n
        return b
    return {"salt": read("salt"), "iv": read("iv"),
            "ct":   read("ct"),   "sig": read("sig")}


def _xim_parse_raw_frames(buf: bytes) -> list:
    """[ {guid, kind, data} ]"""
    out = []
    off = 0
    while off < len(buf):
        if off + 24 > len(buf):
            raise ValueError("Truncated frame")
        guid = buf[off:off+16]; off += 16
        (kind,) = struct.unpack_from("<i", buf, off); off += 4
        (n,)    = struct.unpack_from("<i", buf, off); off += 4
        if n < 0 or off + n > len(buf):
            raise ValueError("Bad frame length")
        data = buf[off:off + n]; off += n
        out.append({"guid": guid, "kind": kind, "data": data})
    return out


def _xim_minify_mmm(image: bytes) -> bytes:
    """Replay of M2M_XIM_Minificator from xim-minifier-x64.dll.

    Mirrors the JS port (_ximMinifyMmm). Header is 0x3DC bytes, section
    table at 0x1DC with 16 entries × 0x20 stride. Each entry's flags
    decide if it's copied; the XOR trailer is appended.
    """
    if image[:3] != b"MMM":
        raise ValueError("Minifier input is not MMM-format")
    HEADER_LEN   = 0x3DC
    TABLE_OFF    = 0x1DC
    ENTRY_STRIDE = 0x20
    N            = 16
    in_buf = bytearray(image)

    needed = HEADER_LEN
    for i in range(N):
        e = TABLE_OFF + i * ENTRY_STRIDE
        if struct.unpack_from("<H", in_buf, e)[0] == 0:
            continue
        if struct.unpack_from("<I", in_buf, e + 0xC)[0] == 1:
            continue
        needed += struct.unpack_from("<I", in_buf, e + 0x8)[0]
    needed += 8  # trailer

    out = bytearray(needed)
    out[:HEADER_LEN] = in_buf[:HEADER_LEN]

    # XOR-of-header: zero out 0x13C, fold the whole header word-by-word,
    # then restore 0x13C; this is what the .NET path does.
    saved_13c = struct.unpack_from("<I", out, 0x13C)[0]
    struct.pack_into("<I", out, 0x13C, 0)
    xor_h = 0
    for p in range(0, HEADER_LEN, 4):
        (w,) = struct.unpack_from("<I", out, p)
        xor_h ^= w
    xor_h &= 0xFFFFFFFF
    struct.pack_into("<I", out, 0x13C, saved_13c)

    # Bump version marker at offset 6 to 0x0900
    struct.pack_into("<H", out, 6, 0x0900)

    out_pos = HEADER_LEN
    for i in range(N):
        e = TABLE_OFF + i * ENTRY_STRIDE
        if struct.unpack_from("<H", out, e)[0] == 0:
            continue
        if struct.unpack_from("<I", out, e + 0xC)[0] == 1:
            continue
        size = struct.unpack_from("<I", out, e + 0x8)[0]
        if size == 0:
            continue
        src_off = struct.unpack_from("<I", out, e + 0x4)[0]
        out[out_pos:out_pos + size] = in_buf[HEADER_LEN + src_off : HEADER_LEN + src_off + size]
        struct.pack_into("<I", out, e + 0x10, out_pos)
        out_pos += size

    xor_s = xor_h
    for p in range(HEADER_LEN, out_pos, 4):
        (w,) = struct.unpack_from("<I", out, p)
        xor_s = (xor_s ^ w) & 0xFFFFFFFF
    if (out_pos & 4) == 0:
        xor_s = (~xor_s) & 0xFFFFFFFF
    struct.pack_into("<I", out, out_pos,     xor_h)
    struct.pack_into("<I", out, out_pos + 4, xor_s)
    return bytes(out)


def _xim_read_leb128(buf: bytes, off: int):
    val = 0; shift = 0; n = 0
    while True:
        if off + n >= len(buf): raise ValueError("LEB128 overflow")
        b = buf[off + n]; n += 1
        val |= (b & 0x7F) << shift
        if not (b & 0x80):
            return val, n
        shift += 7
        if shift > 35:
            raise ValueError("LEB128 too long")


def _xim_read_string(buf: bytes, off: int):
    length, consumed = _xim_read_leb128(buf, off)
    start = off + consumed
    if start + length > len(buf):
        raise ValueError("String overflows frame")
    return buf[start:start + length].decode("utf-8"), consumed + length


def _xim_parse_set(data: bytes) -> set:
    if len(data) < 4: return set()
    (count,) = struct.unpack_from("<i", data, 0)
    out = set(); off = 4
    for _ in range(count):
        s, consumed = _xim_read_string(data, off)
        out.add(s); off += consumed
    return out


def _xim_parse_uint32(data: bytes) -> Optional[int]:
    if len(data) < 4: return None
    return struct.unpack_from("<I", data, 0)[0]


def _xim_parse_version(data: bytes) -> Optional[dict]:
    if len(data) < 16: return None
    mj, mn, bd, rv = struct.unpack_from("<iiii", data, 0)
    return {"major": mj, "minor": mn, "build": bd, "rev": rv}


def _xim_guid(last_byte: int) -> bytes:
    return bytes(15) + bytes([last_byte])


def xim_check_compatibility(parsed: dict, device_info: dict):
    """Host-side compatibility checks. Direct port of `_ximCheckCompatibility`
    in index.html (which is itself a port of IsImeiAllowed / IsFirmwareAllowed
    / IsFirmwareCompatible from the decompiled Teltonika Configurator).

    Returns (ok: bool, blockers: list[str]).

    device_info keys:
      imei            — device's IMEI string
      fw_version_text — current FW like "04.00.00"  (case-insensitive match)
      device_type     — 'FMB' triggers the DeviceId / SpecId gate
    """
    blockers = []
    devs = parsed.get("devices") or set()
    if devs:
        if not device_info.get("imei") or device_info["imei"] not in devs:
            blockers.append(
                "This firmware is locked to specific IMEIs and your device IMEI "
                f"({device_info.get('imei') or 'unknown'}) is not on the list.")
    fws = parsed.get("firmwares") or set()
    if fws:
        cur = (device_info.get("fw_version_text") or "").lower()
        if not any(str(f).lower() == cur for f in fws):
            blockers.append(
                "This firmware can only be installed when the device is already on "
                + " / ".join(sorted(fws))
                + f". Your device is on {device_info.get('fw_version_text') or 'unknown'}.")
    if (device_info.get("device_type") or "FMB") == "FMB":
        did = parsed.get("device_id") or 0
        if did == 0:
            blockers.append("Container has no Device ID — refusing to flash.")
        else:
            parts = [int(p or 0) for p in
                     (device_info.get("fw_version_text") or "0.0.0").split(".")]
            is_pre_v30 = (parts[0] if parts else 0) < 3
            spec_id = parsed.get("spec_id") or 0
            if is_pre_v30 and did == 1 and spec_id != 0:
                blockers.append("This firmware needs a device on FW 3.0 or newer "
                                "for DeviceId=1 + non-zero SpecId.")
    return (len(blockers) == 0, blockers)


def xim_parse(xim_bytes: bytes) -> dict:
    """Decrypt, decompress, parse every metadata frame in a TLTF .xim file.

    Returns dict with keys: image (bytes), device_id, spec_id, version,
    devices (IMEI whitelist), firmwares (FW whitelist), properties.
    """
    parts = _xim_parse_tltf(xim_bytes)
    key = hashlib.pbkdf2_hmac("sha1", _XIM_PASSWORD, parts["salt"], 1000, 32)
    plain = _xim_aes_cbc_decrypt_iso10126(key, parts["iv"], parts["ct"])
    decompressed = gzip.decompress(plain)
    frames = _xim_parse_raw_frames(decompressed)

    def find(kind, last_byte):
        guid = _xim_guid(last_byte)
        for f in frames:
            if f["kind"] == kind and f["guid"] == guid:
                return f
        return None

    img_frame = find(3, 0x04)
    if not img_frame:
        raise ValueError("Image frame not found in TLTF container")

    dev_f  = find(2, 0x01)
    spec_f = find(2, 0x06)
    ver_f  = next((f for f in frames if f["kind"] == 1), None)
    devs_f = find(4, 0x02)
    fws_f  = find(4, 0x03)

    return {
        "image":     _xim_minify_mmm(img_frame["data"]),
        "device_id": _xim_parse_uint32(dev_f["data"])  if dev_f  else None,
        "spec_id":   _xim_parse_uint32(spec_f["data"]) if spec_f else None,
        "version":   _xim_parse_version(ver_f["data"]) if ver_f  else None,
        "devices":   _xim_parse_set(devs_f["data"])    if devs_f else set(),
        "firmwares": _xim_parse_set(fws_f["data"])     if fws_f  else set(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# YMODEM-1K — packet builder + CRC16/XMODEM (decompiled-match)
# ─────────────────────────────────────────────────────────────────────────────
FW_SOH = 0x01; FW_STX = 0x02; FW_EOT = 0x04; FW_ACK = 0x06
FW_NAK = 0x15; FW_CAN = 0x18; FW_SUB = 0x1A; FW_C   = 0x43

# CRC16/XMODEM lookup table (port of Teltonika.FirmwareUpdater.CRC16.Table)
FW_CRC16_TABLE = [
    0,4129,8258,12387,16516,20645,24774,28903,33032,37161,41290,45419,49548,53677,57806,61935,
    4657,528,12915,8786,21173,17044,29431,25302,37689,33560,45947,41818,54205,50076,62463,58334,
    9314,13379,1056,5121,25830,29895,17572,21637,42346,46411,34088,38153,58862,62927,50604,54669,
    13907,9842,5649,1584,30423,26358,22165,18100,46939,42874,38681,34616,63455,59390,55197,51132,
    18628,22757,26758,30887,2112,6241,10242,14371,51660,55789,59790,63919,35144,39273,43274,47403,
    23285,19156,31415,27286,6769,2640,14899,10770,56317,52188,64447,60318,39801,35672,47931,43802,
    27814,31879,19684,23749,11298,15363,3168,7233,60846,64911,52716,56781,44330,48395,36200,40265,
    32407,28342,24277,20212,15891,11826,7761,3696,65439,61374,57309,53244,48923,44858,40793,36728,
    37256,33193,45514,41451,53516,49453,61774,57711,4224,161,12482,8419,20484,16421,28742,24679,
    33721,37784,41979,46042,49981,54044,58239,62302,689,4752,8947,13010,16949,21012,25207,29270,
    46570,42443,38312,34185,62830,58703,54572,50445,13538,9411,5280,1153,29798,25671,21540,17413,
    42971,47098,34713,38840,59231,63358,50973,55100,9939,14066,1681,5808,26199,30326,17941,22068,
    55628,51565,63758,59695,39368,35305,47498,43435,22596,18533,30726,26663,6336,2273,14466,10403,
    52093,56156,60223,64286,35833,39896,43963,48026,19061,23124,27191,31254,2801,6864,10931,14994,
    64814,60687,56684,52557,48554,44427,40424,36297,31782,27655,23652,19525,15522,11395,7392,3265,
    61215,65342,53085,57212,44955,49082,36825,40952,28183,32310,20053,24180,11923,16050,3793,7920,
]


def fw_crc16(data: bytes) -> int:
    crc = 0
    for b in data:
        crc = ((crc << 8) ^ FW_CRC16_TABLE[((crc >> 8) ^ b) & 0xFF]) & 0xFFFF
    return crc


def fw_build_packet(seq: int, payload: bytes) -> bytes:
    """[SOH|STX | seq | 0xFF-seq | payload | crc-hi | crc-lo]"""
    if len(payload) == 128:
        head = FW_SOH
    elif len(payload) == 1024:
        head = FW_STX
    else:
        raise ValueError(f"payload must be 128 or 1024 bytes, got {len(payload)}")
    out = bytearray(3 + len(payload) + 2)
    out[0] = head
    out[1] = seq & 0xFF
    out[2] = (0xFF - (seq & 0xFF)) & 0xFF
    out[3:3 + len(payload)] = payload
    crc = fw_crc16(payload)
    out[-2] = (crc >> 8) & 0xFF
    out[-1] = crc & 0xFF
    return bytes(out)


def fw_build_header_packet(filename: str, size: int) -> bytes:
    txt = filename.encode("utf-8") + b"\x00" + str(size).encode("ascii")
    payload = bytearray(128)
    payload[:min(len(txt), 128)] = txt[:128]
    return fw_build_packet(0, bytes(payload))


def fw_header_filename(original: str) -> str:
    """`xxx.xim` → `xxx.CFW` (preserve case of base, upper-case ext)."""
    base = re.sub(r"\.[^./]+$", "", original or "firmware")
    return base + ".CFW"


def fw_build_data_packets(image: bytes):
    BLOCK = 1024
    n_blocks = (len(image) + BLOCK - 1) // BLOCK
    padded = bytearray(n_blocks * BLOCK)
    padded[:len(image)] = image
    for i in range(len(image), len(padded)):
        padded[i] = FW_SUB
    for b in range(n_blocks):
        seq = (b + 1) & 0xFF
        yield fw_build_packet(seq, bytes(padded[b * BLOCK : (b + 1) * BLOCK]))


def fw_build_null_packet() -> bytes:
    return fw_build_packet(0, bytes(128))


# ─────────────────────────────────────────────────────────────────────────────
# Logger — per-device file log with timestamped TX/RX events
# ─────────────────────────────────────────────────────────────────────────────

LOG_MODE_NONE     = "none"
LOG_MODE_BRIEF    = "brief"     # ASCII command lines + events + errors
LOG_MODE_DETAILED = "detailed"  # every TX/RX byte, hex + ASCII


class Logger:
    """Writes timestamped events to a file. Modes:
      - none     : do nothing
      - brief    : log ASCII command lines, key events, and errors only
      - detailed : log every TX/RX (hex + printable ASCII)

    All file writes go through one mutex so multiple workers (future) and
    the serial RX thread can't tear lines.
    """
    _lock = threading.Lock()

    def __init__(self, path: Optional[str], mode: str):
        self.path = path
        self.mode = mode if mode in (LOG_MODE_BRIEF, LOG_MODE_DETAILED) else LOG_MODE_NONE
        self._fh = None
        if self.path and self.mode != LOG_MODE_NONE:
            try:
                Path(self.path).parent.mkdir(parents=True, exist_ok=True)
                self._fh = open(self.path, "a", encoding="utf-8", buffering=1)
                self._write(f"==== session opened ({self.mode}) ====")
            except Exception:
                self._fh = None  # silently disable on IO error

    @staticmethod
    def _ts() -> str:
        return datetime.now().strftime("%H:%M:%S.") + f"{datetime.now().microsecond // 1000:03d}"

    def _write(self, line: str):
        if not self._fh: return
        with Logger._lock:
            try:
                self._fh.write(f"[{self._ts()}] {line}\n")
            except Exception:
                pass

    def close(self):
        if self._fh:
            try:
                self._write("==== session closed ====")
                self._fh.close()
            except Exception:
                pass
            self._fh = None

    # ── Helpers used by SerialLink / worker ──
    def event(self, msg: str):
        # Events always log (brief + detailed)
        self._write(f"--- {msg}")

    def error(self, msg: str):
        self._write(f"!!! ERROR: {msg}")

    def tx(self, data: bytes, *, ymodem: bool = False):
        if self.mode == LOG_MODE_NONE: return
        if self.mode == LOG_MODE_BRIEF:
            # Brief: skip YMODEM packets (too noisy); log ASCII commands only
            if ymodem: return
            try:
                s = data.decode("ascii", errors="replace").replace("\r", "\\r")
                self._write(f"TX  {s}")
            except Exception: pass
            return
        # Detailed: full hex + ascii
        self._write(f"TX  {_fmt_bytes(data)}")

    def rx(self, data: bytes, *, ymodem: bool = False):
        if self.mode == LOG_MODE_NONE: return
        if not data: return
        if self.mode == LOG_MODE_BRIEF:
            if ymodem: return
            try:
                s = data.decode("ascii", errors="replace").replace("\r", "\\r")
                self._write(f"RX  {s}")
            except Exception: pass
            return
        self._write(f"RX  {_fmt_bytes(data)}")


def _fmt_bytes(b: bytes, max_len: int = 64) -> str:
    """Compact hex + ascii view for the log. Truncates very long lines."""
    if len(b) > max_len:
        head = b[:max_len]
        suffix = f" …(+{len(b) - max_len}B)"
    else:
        head = b
        suffix = ""
    hex_part = " ".join(f"{x:02X}" for x in head)
    ascii_part = "".join(chr(x) if 32 <= x < 127 else "." for x in head)
    return f"{hex_part}  | {ascii_part}{suffix}"


# ─────────────────────────────────────────────────────────────────────────────
# SerialLink — thin wrapper around pyserial with line-oriented helpers
# ─────────────────────────────────────────────────────────────────────────────

class DeviceDisconnected(Exception):
    """Raised by SerialLink when the underlying COM port reports a hard
    failure (cable yanked, USB hub power-cycled, driver loss). Catchable in
    the worker so each step can decide whether to retry, abort, or switch
    to a different recovery flow (firmware mid-flight = power-cycle path)."""
    pass


# Errors raised by pyserial when the port disappears. We translate them
# to DeviceDisconnected at the SerialLink boundary so callers never have
# to know about pyserial internals.
_DISCONNECT_ERRORS = (serial.SerialException, OSError, PermissionError)


class SerialLink:
    """Open/close wrapper with read_until + drain helpers. If a logger is
    attached every TX/RX gets timestamped to the log file.

    Every read/write operation is wrapped to detect a hard disconnect and
    raise DeviceDisconnected. Callers can then retry the step or fall
    through to a higher-level recovery."""

    def __init__(self, port: str, baud: int = BAUD, logger: Optional[Logger] = None):
        self.port = port
        self.baud = baud
        self._s: Optional[serial.Serial] = None
        self.logger = logger or Logger(None, LOG_MODE_NONE)
        # YMODEM-mode flag: while set, TX/RX is treated as "ymodem" and is
        # suppressed in brief log mode (so we don't drown the log in 2992
        # 1KB packets). Worker sets/unsets around the flash inner loop.
        self.ymodem_mode = False

    def open(self):
        if self._s and self._s.is_open:
            return
        try:
            self._s = serial.Serial(self.port, self.baud,
                                    timeout=0.05, write_timeout=2.0)
        except _DISCONNECT_ERRORS as e:
            raise DeviceDisconnected(f"open({self.port}) failed: {e}")

    def close(self):
        if self._s:
            try: self._s.close()
            except Exception: pass
            self._s = None

    def is_open(self) -> bool:
        return bool(self._s and self._s.is_open)

    # ── internal raw I/O wrappers (the ONLY pyserial touchpoints) ───────
    def _raw_read(self, n: int) -> bytes:
        if not self._s:
            raise DeviceDisconnected("port closed")
        try:
            return self._s.read(n)
        except _DISCONNECT_ERRORS as e:
            raise DeviceDisconnected(f"read failed: {e}")

    def _raw_write(self, data: bytes):
        if not self._s:
            raise DeviceDisconnected("port closed")
        try:
            self._s.write(data)
            self._s.flush()
        except _DISCONNECT_ERRORS as e:
            raise DeviceDisconnected(f"write failed: {e}")

    def _raw_in_waiting(self) -> int:
        if not self._s:
            raise DeviceDisconnected("port closed")
        try:
            return self._s.in_waiting
        except _DISCONNECT_ERRORS as e:
            raise DeviceDisconnected(f"in_waiting failed: {e}")

    # ── public I/O (these raise DeviceDisconnected on hard failures) ────
    def write(self, data: bytes):
        self._raw_write(data)
        try: self.logger.tx(data, ymodem=self.ymodem_mode)
        except Exception: pass

    def write_ascii(self, cmd: str):
        """Send a `:foo:bar\\r` ASCII command. Caller passes no trailing CR."""
        self.write(cmd.encode("ascii") + b"\r")

    def drain(self, ms: int = 100):
        if not self._s: return
        deadline = time.monotonic() + ms / 1000.0
        while time.monotonic() < deadline:
            b = self._raw_read(4096)
            if b:
                try: self.logger.rx(b, ymodem=self.ymodem_mode)
                except Exception: pass
            else:
                time.sleep(0.01)

    def read_until(self, terminator: bytes, timeout_s: float) -> bytes:
        deadline = time.monotonic() + timeout_s
        buf = bytearray()
        while time.monotonic() < deadline:
            chunk = self._raw_read(4096)
            if chunk:
                buf.extend(chunk)
                try: self.logger.rx(chunk, ymodem=self.ymodem_mode)
                except Exception: pass
                if terminator in buf:
                    return bytes(buf)
        return bytes(buf)

    def read_for(self, secs: float) -> bytes:
        deadline = time.monotonic() + secs
        buf = bytearray()
        while time.monotonic() < deadline:
            chunk = self._raw_read(4096)
            if chunk:
                buf.extend(chunk)
                try: self.logger.rx(chunk, ymodem=self.ymodem_mode)
                except Exception: pass
            else:
                time.sleep(0.005)
        return bytes(buf)

    def read_byte_match(self, pred: Callable[[int], bool], timeout_s: float) -> Optional[int]:
        """Low-latency single-byte wait. Uses in_waiting polling to avoid
        pyserial's per-read timeout cost."""
        deadline = time.monotonic() + timeout_s
        while time.monotonic() < deadline:
            n = self._raw_in_waiting()
            if n:
                chunk = self._raw_read(n)
                try: self.logger.rx(chunk, ymodem=self.ymodem_mode)
                except Exception: pass
                for b in chunk:
                    if pred(b): return b
            else:
                time.sleep(0.0005)
        return None

    def read_line_match(self, pattern: re.Pattern, timeout_s: float) -> Optional[str]:
        deadline = time.monotonic() + timeout_s
        buf = bytearray()
        while time.monotonic() < deadline:
            chunk = self._raw_read(4096)
            if chunk:
                buf.extend(chunk)
                try: self.logger.rx(chunk, ymodem=self.ymodem_mode)
                except Exception: pass
                txt = buf.decode("ascii", errors="replace")
                for line in txt.splitlines():
                    if pattern.search(line):
                        return line
            else:
                time.sleep(0.005)
        return None


def port_still_present(port: str) -> bool:
    """Backstop for silent disconnects: returns True if `port` is still in
    list_ports.comports(). pyserial sometimes won't error on a dead port
    until the next write; this catches that case before we block on a
    read that will never complete."""
    try:
        return any(p.device == port for p in list_ports.comports())
    except Exception:
        return True  # err on the side of "still here" if we can't tell


# ─────────────────────────────────────────────────────────────────────────────
# Device-identity probe (parses `:cfg_info:?` slot stream)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class DeviceIdentity:
    imei: str = ""
    firmware: str = ""
    config_version: str = ""
    model: str = ""
    spec_id: str = "1"
    hw_version: str = ""

_CFG_INFO_RE = re.compile(r"cfg_info:(\d+):([^\r\n]*)", re.IGNORECASE)


def parse_identity(text: str) -> Optional[DeviceIdentity]:
    slots = {int(m.group(1)): m.group(2).strip()
             for m in _CFG_INFO_RE.finditer(text)}
    if not slots: return None
    ident = DeviceIdentity()
    if 3 in slots and slots[3].isdigit() and len(slots[3]) >= 14:
        ident.imei = slots[3]
    if 0 in slots:
        fw = slots[0]; rev = slots.get(13, "").strip()
        ident.firmware = f"{fw} Rev:{rev}" if rev else fw
    if 1 in slots: ident.config_version = slots[1]
    if 2 in slots:
        raw = slots[2]
        ident.model = MODEL_VARIANT_MAP.get(raw, raw)
        # Family prefixes we recognise. Add new prefixes here when
        # MODEL_VARIANT_MAP grows beyond FM*/MTB.
        _KNOWN_PREFIXES = ("FM", "MTB")
        if not ident.model.startswith(_KNOWN_PREFIXES):
            mm = re.search(r"\b(FM[BMCPTU]\w{1,4}|MTB\d{1,4})\b", raw)
            if mm: ident.model = mm.group(1).upper()
        ident.hw_version = raw
    if 11 in slots: ident.spec_id = slots[11] or "1"
    # Success if we have IMEI or recognised model
    if ident.imei or (ident.model and ident.model.startswith(("FM", "MTB"))):
        return ident
    return None


def probe_port(port: str, timeout_s: float = PROBE_TIMEOUT_S) -> Optional[DeviceIdentity]:
    """Open, silence log spam, ask identity, close. None = not an FMB."""
    link = SerialLink(port)
    try:
        try: link.open()
        except Exception: return None
        try:
            link.write(b".log:0\r"); link.drain(100)
        except Exception: pass
        try:
            link.write(b":cfg_info:?\r")
        except Exception:
            return None
        raw = link.read_for(timeout_s)
        if not raw: return None
        text = raw.decode("ascii", errors="replace")
        return parse_identity(text)
    finally:
        link.close()


# ─────────────────────────────────────────────────────────────────────────────
# Security commands  (sec_status / sec_login / sec_setkey / sec_changekey / sec_removekey)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SecStat:
    is_valid_keyword: int = 0   # 1 = OK
    has_keyword:      int = 0   # 1 = device is secured
    authenticated:    int = 0   # 1 = session is authenticated
    locked:           int = 0   # 1 = no retries left
    retries:          int = 0


def parse_secstat(line: str) -> Optional[SecStat]:
    """`<SECSTAT>1,1,1,0,5` → SecStat. Returns None if no match."""
    m = re.search(r"<SECSTAT>([^\r\n]*)", line)
    if not m: return None
    parts = (m.group(1) + ",,,,,").split(",")
    def i(p):
        try: return int(p)
        except: return 0
    return SecStat(i(parts[0]), i(parts[1]), i(parts[2]), i(parts[3]), i(parts[4]))


def cmd_sec_status(link: SerialLink) -> Optional[SecStat]:
    link.write(b":sec_status\r")
    line = link.read_line_match(re.compile(r"<SECSTAT>"), SECSTAT_TIMEOUT_S)
    return parse_secstat(line) if line else None


def cmd_sec_login(link: SerialLink, keyword: str) -> Optional[SecStat]:
    link.write_ascii(f":sec_login:{keyword}")
    line = link.read_line_match(re.compile(r"<SECSTAT>"), SECSTAT_TIMEOUT_S)
    return parse_secstat(line) if line else None


def cmd_sec_setkey(link: SerialLink, new_kw: str) -> Optional[SecStat]:
    """Set a new keyword on an unsecured device. NEW first, CONFIRM second."""
    link.write_ascii(f":sec_setkey:{new_kw},{new_kw}")
    line = link.read_line_match(re.compile(r"<SECSTAT>"), SECSTAT_TIMEOUT_S)
    return parse_secstat(line) if line else None


def cmd_sec_changekey(link: SerialLink, new_kw: str, current_kw: str) -> Optional[SecStat]:
    """Change keyword on a secured device. NEW first, CURRENT second."""
    link.write_ascii(f":sec_changekey:{new_kw},{current_kw}")
    line = link.read_line_match(re.compile(r"<SECSTAT>"), SECSTAT_TIMEOUT_S)
    return parse_secstat(line) if line else None


def cmd_sec_removekey(link: SerialLink) -> Optional[SecStat]:
    """Remove keyword. Device must already be authenticated."""
    link.write_ascii(":sec_removekey")
    line = link.read_line_match(re.compile(r"<SECSTAT>"), SECSTAT_TIMEOUT_S)
    return parse_secstat(line) if line else None


def cmd_cfg_connect(link: SerialLink, timeout_s: float = 1.5) -> bool:
    """Send :cfg_connect, wait for <CFG_CONNECT>. Configurator does this
    on every fresh session before sec_status; we do the same so the device
    is in the expected configurator-aware mode."""
    link.write_ascii(":cfg_connect")
    line = link.read_line_match(re.compile(r"<CFG_CONNECT>"), timeout_s)
    return line is not None


def cmd_cfg_getcfg(link: SerialLink, progress: Optional[Callable[[int, str], None]] = None,
                   timeout_s: float = 30.0) -> dict:
    """Read the entire device config via :cfg_getcfg.

    Wire format (per the configurator):
        <GET_PARAMS_START>
        pid:value\\r
        pid:value\\r
        …
        <GET_PARAMS_END>:<count>

    Returns { pid_int: value_str }. On timeout returns whatever was parsed
    before the deadline (callers diff against this; partial is fine).
    """
    link.write_ascii(":cfg_getcfg")
    out: dict = {}
    deadline = time.monotonic() + timeout_s
    buf = bytearray()
    seen_start = False
    line_pat = re.compile(r"^(\d+):(.*)$")
    last_progress = 0
    while time.monotonic() < deadline:
        # Use the DeviceDisconnected-aware accessors. A hard disconnect now
        # raises instead of silently returning partial data — the step
        # wrapper catches it and triggers a clean reconnect-retry.
        n = link._raw_in_waiting()
        chunk = link._raw_read(n) if n else b""
        if not chunk:
            time.sleep(0.005); continue
        try: link.logger.rx(chunk)
        except Exception: pass
        buf.extend(chunk)
        # Split on \r and \n
        while True:
            idx = -1
            for sep in (b"\r", b"\n"):
                p = buf.find(sep)
                if p >= 0 and (idx < 0 or p < idx):
                    idx = p
            if idx < 0: break
            raw_line = bytes(buf[:idx]); del buf[:idx + 1]
            line = raw_line.decode("latin1", errors="replace").strip()
            if not line: continue
            if "<GET_PARAMS_START>" in line:
                seen_start = True; continue
            if line.startswith("<GET_PARAMS_END>"):
                return out
            if not seen_start: continue
            m = line_pat.match(line)
            if m:
                try:
                    out[int(m.group(1))] = m.group(2)
                except Exception: pass
                if progress and len(out) // 500 > last_progress:
                    last_progress = len(out) // 500
                    progress(50, f"Loaded {len(out)} parameters from device…")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Config file parse + upload (batched cfg_setparam + cfg_save)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class CfgFile:
    header: dict = field(default_factory=dict)   # ConfigurationVersion, HwVersion, FmType, SpecId, Title
    pids: dict   = field(default_factory=dict)   # int → str


def parse_cfg_bytes(raw: bytes) -> CfgFile:
    """`.cfg` is gzipped `Key:Value;Key:Value;...;<pid>:<val>;...`.

    Header keys (non-numeric): ConfigurationVersion, HwVersion, Title,
    FmType, SpecId. Body is `<pid>:<value>` semicolon-separated.
    """
    # Try gunzip; fall back to raw if it isn't gzipped.
    try:
        text_bytes = gzip.decompress(raw)
    except OSError:
        text_bytes = raw
    text = text_bytes.decode("latin1", errors="replace")
    cfg = CfgFile()
    for part in text.split(";"):
        if ":" not in part: continue
        k, _, v = part.partition(":")
        k = k.strip(); v = v.strip()
        if k.isdigit():
            cfg.pids[int(k)] = v
        elif k:
            cfg.header[k] = v
    return cfg


def _cfg_val_norm(v) -> str:
    """Normalise a cfg value for diffing — matches `_cfgValNorm` in index.html.

    Treats undefined / empty / whitespace as equal; strips leading zeros from
    integer-looking strings; trims trailing zeros after the decimal point on
    floats. Without this, a value the device omitted (1600:'') counts as
    "different" from the file's literal empty entry, producing huge dirty
    lists.
    """
    if v is None: return ""
    s = str(v).strip()
    if s == "": return ""
    # Pure integer (with possible leading sign): strip leading zeros.
    if re.match(r"^-?\d+$", s):
        sign = "-" if s.startswith("-") else ""
        digits = s.lstrip("-").lstrip("0") or "0"
        return sign + digits
    # Float-looking: trim trailing zeros after the decimal point.
    if re.match(r"^-?\d+\.\d+$", s):
        sign = "-" if s.startswith("-") else ""
        whole, _, frac = s.lstrip("-").partition(".")
        frac = frac.rstrip("0")
        if not frac: return sign + (whole.lstrip("0") or "0")
        return sign + (whole.lstrip("0") or "0") + "." + frac
    return s


def _build_setparam_batches(pairs, max_bytes: int = SETPARAM_BATCH_B):
    """[(pid, val), ...]  →  [(body_text, [pid_list]), ...] sized for one wire line."""
    out = []
    buf = ""; batch_pids = []
    prefix_len = len(":cfg_setparam:") + 1
    for pid, val in pairs:
        piece = (";" if buf else "") + f"{pid}:{val if val is not None else ''}"
        if buf and (prefix_len + len(buf) + len(piece)) > max_bytes:
            out.append((buf, batch_pids))
            buf = f"{pid}:{val if val is not None else ''}"
            batch_pids = [pid]
        else:
            buf += piece
            batch_pids.append(pid)
    if buf: out.append((buf, batch_pids))
    return out


def upload_cfg(link: SerialLink, cfg: CfgFile,
               progress: Callable[[int, str], None]) -> tuple[int, int, int, list]:
    """Diff-then-save (matches `_handleCfgFile` + `saveDirtyParams` in the
    web configurator).

    Step 1: read device cfg via :cfg_getcfg (so we know its current state).
    Step 2: walk the file PIDs and select only those that differ from device
            and aren't in the setparam blacklist (e.g. PID 10 = keyword).
    Step 3: send only the modified PIDs via batched :cfg_setparam, then :cfg_save.

    Returns (saved, matched, modified, failed_pids):
      matched   — number of file PIDs whose values already equal the device
      modified  — number of file PIDs we actually attempted to write
      saved     — number of those modified PIDs the device accepted
      failed_pids — PIDs that appeared in a batch the device only partially
                    accepted (marked conservatively — some MAY have written)
    """
    # ── Read device cfg, retrying if the device isn't ready yet ─────────
    # After a firmware update the device's cfg table can be empty for a few
    # seconds while it populates from defaults. If we charge in with cfg_getcfg
    # too early the device returns just `<GET_PARAMS_END>:0` and we mistake
    # every file PID as dirty — sending ~11k params instead of ~100.
    # Wait + retry until we get a non-empty response, or give up after a
    # few attempts (then proceed with all-dirty as a last resort).
    CFG_READY_RETRIES = 4
    CFG_READY_DELAY_S = 3.0
    device_cfg = {}
    for attempt in range(1, CFG_READY_RETRIES + 1):
        progress(2, f"Reading current configuration from device (attempt {attempt})…")
        device_cfg = cmd_cfg_getcfg(link, progress=progress)
        link.logger.event(f"cfg_getcfg attempt {attempt}: loaded {len(device_cfg)} PIDs")
        if device_cfg:
            break
        if attempt < CFG_READY_RETRIES:
            progress(2, f"Device not ready yet (0 PIDs returned). "
                        f"Waiting {CFG_READY_DELAY_S:.0f}s then retrying…")
            link.logger.event(f"cfg_getcfg returned 0 PIDs; waiting "
                              f"{CFG_READY_DELAY_S}s before retry")
            time.sleep(CFG_READY_DELAY_S)
    if not device_cfg:
        link.logger.event(f"cfg_getcfg still empty after {CFG_READY_RETRIES} "
                          "attempts — falling through to all-dirty write")

    # PIDs the device will NEVER accept via :cfg_setparam — they're in the
    # security block and can only be changed through :sec_setkey /
    # :sec_changekey / :sec_removekey. cfg_getcfg still reports them, so if
    # we don't filter here every future run flags them as "dirty" and the
    # setparam always fails, producing a phantom error in the grid.
    CFG_SETPARAM_BLACKLIST = {10}   # PID 10 = keyword slot
    # Build the "modified" list — only PIDs that actually differ from device
    # AND that we're allowed to write via :cfg_setparam.
    modified = []
    matched = 0
    for pid, val in sorted(cfg.pids.items()):
        dev_val = device_cfg.get(pid)
        if _cfg_val_norm(dev_val) == _cfg_val_norm(val):
            matched += 1
            continue
        if pid in CFG_SETPARAM_BLACKLIST:
            # Silently absorb — these are not real modifications from the
            # operator's perspective (they can never be written this way).
            matched += 1
            continue
        modified.append((pid, val))
    link.logger.event(f"diff: {len(modified)} of {len(cfg.pids)} file PIDs modified vs device")

    if not modified:
        progress(100, "Device already matches the file — no parameters to write.")
        return (0, matched, 0, [])

    batches = _build_setparam_batches(modified)
    saved = 0
    failed = []
    total = len(batches)
    for i, (body, pids) in enumerate(batches):
        pct = 5 + int((i / max(total, 1)) * 85)
        progress(pct, f"Writing batch {i + 1}/{total} ({len(pids)} parameters)")
        link.write_ascii(":cfg_setparam:" + body)
        line = link.read_line_match(re.compile(r"<SETPARAM_RESULT>:\d+|^OK"), SETPARAM_TIMEOUT_S)
        if not line:
            failed.extend(pids); continue
        m = re.search(r"<SETPARAM_RESULT>:(\d+)", line)
        if m:
            n = int(m.group(1))
            if n == len(pids):
                saved += n
            else:
                # Partial — device accepted only `n` of the batch. We don't
                # know which N succeeded; mark all in this batch as failed
                # to be conservative. NOTE: this is NOT a real error — the
                # device just declined some PIDs; the row stays green.
                failed.extend(pids)
        elif line.strip() == "OK":
            saved += len(pids)
        else:
            failed.extend(pids)
    progress(92, "Committing :cfg_save…")
    link.write_ascii(":cfg_save")
    link.read_line_match(re.compile(r"<\w+>:\d+"), CFG_SAVE_TIMEOUT_S)
    progress(100, f"Saved {saved} of {len(modified)} modified parameters "
                  f"(file had {len(cfg.pids)}, device matched {matched})")
    return (saved, matched, len(modified), failed)


# ─────────────────────────────────────────────────────────────────────────────
# Firmware handshake + YMODEM-1K flash loop  (decompiled-match)
# ─────────────────────────────────────────────────────────────────────────────

def _fw_send_ascii(link: SerialLink, cmd: str):
    link.write(cmd.encode("ascii") + b"\r")


def _fw_wait_line(link: SerialLink, pattern: re.Pattern, timeout_s: float) -> Optional[str]:
    return link.read_line_match(pattern, timeout_s)


def fw_handshake(link: SerialLink, container: dict, progress: Callable[[str], None]):
    """Pre-flight: querying boot-mode, IMEI/info checks, UPDQUERY,
    pause, second device-info, '.run_uart_boot_mode' (POINT OF NO RETURN),
    wait for first 'C' from bootloader. Raises on any failure."""
    progress("Querying boot-mode availability…")
    _fw_send_ascii(link, ".run_uart_boot_mode?")
    fwst = _fw_wait_line(link, re.compile(r"^<UPDATEFWST>:(-?\d+)"), 5.0)
    if not fwst:
        raise RuntimeError("No response to .run_uart_boot_mode?")
    code = int(re.search(r"<UPDATEFWST>:(-?\d+)", fwst).group(1))
    if code != 0:
        raise RuntimeError(f"Device refused boot mode (UPDATEFWST code {code})")

    progress("Reading device info…")
    _fw_send_ascii(link, ".tst_device_info")
    if not _fw_wait_line(link, re.compile(r"^tst:fm_device_inf:"), 5.0):
        raise RuntimeError("No response to .tst_device_info")

    progress("Reading IMEI…")
    _fw_send_ascii(link, ".imei")
    if not _fw_wait_line(link, re.compile(r"^IMEI:\d+"), 5.0):
        raise RuntimeError("No response to .imei")

    progress("Verifying firmware compatibility…")
    did = container.get("device_id", -1) or -1
    ver = container.get("version") or {"major": 0, "minor": 0, "build": 0}
    sid = container.get("spec_id", 0) or 0
    fmt2 = lambda n: str(max(0, int(n))).zfill(2)
    q = f":cfg_updquery:{did},{fmt2(ver['major'])}.{fmt2(ver['minor'])}.{fmt2(ver['build'])}"
    if sid: q += f",{sid}"
    _fw_send_ascii(link, q)
    updq = _fw_wait_line(link, re.compile(r"^<UPDQUERY>:(-?\d+)"), 5.0)
    if not updq:
        raise RuntimeError("No response to UPDQUERY")
    code = int(re.search(r"<UPDQUERY>:(-?\d+)", updq).group(1))
    msgs = {0: "Compatible", 1: "Wrong device ID", 2: "Wrong firmware version",
            3: "Wrong spec ID", 4: "Validation error"}
    if code != 0:
        raise RuntimeError(f"Firmware not compatible: {msgs.get(code, f'code {code}')}")

    progress("Settling…")
    time.sleep(1.0)

    progress("Re-verifying device info…")
    _fw_send_ascii(link, ".tst_device_info")
    if not _fw_wait_line(link, re.compile(r"^tst:fm_device_inf:"), 5.0):
        raise RuntimeError("No response to second .tst_device_info")

    progress("Re-verifying IMEI…")
    _fw_send_ascii(link, ".imei")
    if not _fw_wait_line(link, re.compile(r"^IMEI:\d+"), 5.0):
        raise RuntimeError("No response to second .imei")

    # ========== POINT OF NO RETURN ==========
    progress("Entering bootloader mode…")
    _fw_send_ascii(link, ".run_uart_boot_mode")
    if not _fw_wait_line(link, re.compile(r"^FW Update start"), 8.0):
        raise RuntimeError("Bootloader didn't acknowledge boot-mode entry")

    progress("Waiting for bootloader 'C'…")
    if link.read_byte_match(lambda b: b == FW_C, 15.0) is None:
        raise RuntimeError("Bootloader 'C' not received within 15s")


def _fw_send_with_retries(link: SerialLink, packet: bytes,
                          wait_for: str, max_retries: int,
                          per_try_timeout_s: float,
                          on_retry: Callable[[int, int], None] = None):
    """Send packet, wait for ACK. Optional follow-up C drain for the header.

    Port of `fwSendPacketAwait` in index.html — exact same retry semantics:
      - 1 + max_retries total attempts
      - Drain stale rx before each TX so we don't pick up an old ACK
      - Wait for one of ACK / NAK / CAN
      - ACK → optionally drain a following C → done
      - CAN → abort with explicit error
      - NAK or timeout → retry (next attempt invokes on_retry callback)
    """
    for attempt in range(max_retries + 1):
        # Drain stale rx (any leftover bytes from a prior step). Goes
        # through the DeviceDisconnected-aware path so a yanked cable here
        # surfaces as DeviceDisconnected up to the caller — pyserial alone
        # would just have silently returned 0 bytes.
        try:
            n = link._raw_in_waiting()
            if n: link._raw_read(n)
        except DeviceDisconnected:
            raise   # propagate to the firmware step's handler
        except Exception:
            pass
        if attempt > 0 and on_retry:
            on_retry(attempt, max_retries)

        link.write(packet)
        b = link.read_byte_match(lambda x: x in (FW_ACK, FW_NAK, FW_CAN),
                                 per_try_timeout_s)
        if b == FW_ACK:
            if wait_for == "C+ACK":
                # Drain the following C the bootloader emits to start the next
                # packet. Don't fail if it's silent — some firmwares skip it.
                link.read_byte_match(lambda x: x == FW_C, per_try_timeout_s)
            return
        if b == FW_CAN:
            raise RuntimeError("Device sent CAN — bootloader aborted the transfer")
        # NAK or None → retry
    raise RuntimeError(f"Packet not acknowledged after {max_retries + 1} attempts")


def _fw_data_phase(link: SerialLink, container: dict, file_name: str,
                   progress: Callable[[int, str], None],
                   cancel_check: Callable[[], bool]):
    """Header → data packets → EOT → null. Called AFTER fw_handshake has
    successfully entered bootloader mode. Disconnect here means the device
    is in bootloader limbo — caller must handle that case (no clean retry)."""
    image = container["image"]
    progress(10, "Sending firmware header…")
    header_name = fw_header_filename(file_name)
    _fw_send_with_retries(link, fw_build_header_packet(header_name, len(image)),
                          "C+ACK", 4, 5.0,
                          lambda a, m: progress(10, f"Header retry {a}/{m}…"))
    packets = list(fw_build_data_packets(image))
    total = len(packets)
    link.ymodem_mode = True
    try:
        for i, pkt in enumerate(packets):
            if cancel_check():
                raise RuntimeError("user cancelled (no recovery — device will time out)")
            pct = 10 + int(((i + 1) / total) * 80)
            progress(pct, f"Flashing packet {i + 1} / {total}")
            _fw_send_with_retries(link, pkt, "ACK", 4, 5.0,
                                  lambda a, m, j=i: progress(pct, f"Packet {j+1}/{total} — retry {a}/{m}…"))
    finally:
        link.ymodem_mode = False
    progress(92, "End-of-transmission…")
    link.write(bytes([FW_EOT]))
    link.read_byte_match(lambda b: b == FW_ACK, 10.0)
    progress(95, "Final packet…")
    link.write(fw_build_null_packet())
    link.read_byte_match(lambda b: b == FW_ACK, 5.0)
    progress(100, "Transfer complete — device flashing (~3 min reboot)")


def flash_firmware(link: SerialLink, container: dict, file_name: str,
                   progress: Callable[[int, str], None],
                   cancel_check: Callable[[], bool]):
    """Full firmware update path (kept for back-compat / single-shot callers).
    Worker uses fw_handshake + _fw_data_phase directly so it can detect the
    bootloader boundary for disconnect handling."""
    progress(2, "Pre-flight handshake…")
    fw_handshake(link, container, lambda label: progress(3, label))
    _fw_data_phase(link, container, file_name, progress, cancel_check)


# ─────────────────────────────────────────────────────────────────────────────
# Port detection — VID/PID filter + 2-port pairing
# ─────────────────────────────────────────────────────────────────────────────

def list_teltonika_ports() -> list:
    """Return list_ports entries whose VID/PID matches a Teltonika filter.

    FMB devices enumerate TWO ports per device (debug + config). Each
    candidate is returned as-is; pairing-to-config-port happens via probe.
    """
    out = []
    for p in list_ports.comports():
        vid = f"{p.vid:04X}" if p.vid is not None else ""
        pid = f"{p.pid:04X}" if p.pid is not None else ""
        for fvid, fpid in USB_DEVICE_FILTERS:
            if vid.upper() == fvid and pid.upper() == fpid:
                out.append(p)
                break
    return out


def find_config_port() -> Optional[tuple]:
    """Scan known-Teltonika ports, probe each. Return (port_name, identity)
    of the first port that answers like a config port (responds to :cfg_info:?).
    Returns None if no responsive Teltonika is found right now."""
    for p in list_teltonika_ports():
        ident = probe_port(p.device)
        if ident:
            return (p.device, ident)
    return None


def all_current_ports() -> set:
    """Snapshot of all current COM port names (incl. non-Teltonika).
    Used by the port-change watcher."""
    return {p.device for p in list_ports.comports()}


# ─────────────────────────────────────────────────────────────────────────────
# Batch options + per-device worker
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class BatchOptions:
    do_firmware:     bool = False
    firmware_path:   str  = ""
    do_config:       bool = False
    config_path:     str  = ""
    keyword_action:  str  = "none"  # "set" | "remove" | "none"
    new_keyword:     str  = ""      # for "set"
    current_keyword: str  = ""      # auth + Remove
    # Single shared log file; always full (detailed) log; truncated by the
    # app on Start, appended-to by each device's Logger.
    log_path:        str  = ""      # absolute path to batch_configurator.log


@dataclass
class DeviceResult:
    timestamp:    str = ""
    port:         str = ""
    imei:         str = ""
    model:        str = ""
    firmware:     str = ""
    config_file:  str = ""
    keyword:      str = ""   # final keyword state (raw value or "")
    # Free-form summary shown in the grid's Remarks column. Contains an
    # informational summary of what the batch did per device (e.g.
    # "Config: 8873 params matched, 2 of 3 modified written"). Empty
    # unless the worker had something worth mentioning.
    remarks:      str = ""
    # True when a genuine failure occurred (disconnect / auth fail /
    # firmware fail / etc.). Config-side partial writes are NOT errors —
    # they're just informational. Grid uses this flag to decide red vs
    # green row.
    has_error:    bool = False


class CancelRequested(Exception):
    pass


class DeviceWorker:
    """Runs the full enabled-step sequence against ONE connected device.

    All UI updates happen via the `ui` callback, which the main thread
    routes to the progress modal. `cancel_evt` is checked at coarse
    boundaries (between steps and between firmware packets) — never mid-
    packet, never mid-batch (we let the current step finish first).
    """
    def __init__(self, port: str, ident: DeviceIdentity, opts: BatchOptions,
                 ui_progress: Callable[[str, int, str], None],
                 cancel_evt: threading.Event,
                 skip_firmware_evt: threading.Event,
                 ask_keyword: Callable[[str], Optional[str]]):
        self.port = port
        self.ident = ident
        self.opts = opts
        self.ui = ui_progress      # (step_name, pct 0..100, detail) → None
        self.cancel_evt = cancel_evt
        self.skip_firmware_evt = skip_firmware_evt
        # ask_keyword(imei) → string (new keyword) or None (=cancel this device).
        # Used when the form's Current Keyword is wrong/missing for a secured
        # device — we ask the user for a per-device replacement so one bad
        # device doesn't destroy the whole batch.
        self.ask_keyword = ask_keyword
        # Remarks accumulate across steps. `_has_error` is flipped only for
        # genuine failures; informational messages (config summaries,
        # "skipped" notes) get appended without flipping the flag so the
        # grid row stays green.
        self.errors = []          # legacy attr name; holds remark strings
        self._has_error = False
        # The device's CURRENT keyword (post-batch state). Empty means the
        # device has no keyword set right now. This is what gets written to
        # the grid's Keyword column / CSV — not status text like "Removed".
        self._device_keyword: str = ""
        # Logger handle — populated for real at the top of run() so the
        # disconnect-recovery plumbing (which can fire on any step) has a
        # place to write events without crashing on a missing attribute.
        self._logger: Logger = Logger(None, LOG_MODE_NONE)

    def _check_cancel_between_steps(self):
        if self.cancel_evt.is_set():
            raise CancelRequested()

    # ── Reconnect plumbing ───────────────────────────────────────────────
    # Outcomes that _wait_for_reconnect returns. Per-step wrappers decide
    # what to do with each:
    #    same       — link is open on a port whose IMEI matches ours; retry
    #    different  — another Teltonika appeared with a different IMEI;
    #                 abort this device, main loop picks up the new one
    #    timeout    — nothing came back within the window; abort this device
    #    cancelled  — user clicked Cancel; abort
    RECONNECT_TIMEOUT_S = 60
    RECONNECT_RETRIES_PER_STEP = 3

    def _wait_for_reconnect(self, link: SerialLink, context: str) -> str:
        """Block until the device returns or we give up. `context` is the
        free-form text shown in the modal (e.g. "during config upload"
        or "during firmware transfer — power-cycle if stuck").
        On 'same', the link is reopened on the (possibly new) port name."""
        try: link.close()
        except Exception: pass
        self.ui("Reconnect", 0, context)
        deadline = time.monotonic() + self.RECONNECT_TIMEOUT_S
        last_secs_shown = -1
        while time.monotonic() < deadline:
            if self.cancel_evt.is_set():
                return "cancelled"
            remaining = int(deadline - time.monotonic())
            if remaining != last_secs_shown:
                last_secs_shown = remaining
                self.ui("Reconnect", remaining, context)
            # Try the same port first.
            if port_still_present(self.port):
                ident = probe_port(self.port)
                if ident:
                    if ident.imei == self.ident.imei:
                        try:
                            link.port = self.port
                            link.open()
                            return "same"
                        except DeviceDisconnected:
                            pass  # transient — keep looping
                    else:
                        return "different"
            # Fall back to scanning all Teltonika ports for our IMEI.
            for p in list_teltonika_ports():
                if self.cancel_evt.is_set(): return "cancelled"
                if p.device == self.port: continue  # already tried above
                ident = probe_port(p.device)
                if not ident: continue
                if ident.imei == self.ident.imei:
                    try:
                        self.port = p.device
                        link.port = p.device
                        link.open()
                        return "same"
                    except DeviceDisconnected:
                        pass
                else:
                    return "different"
            time.sleep(0.5)
        return "timeout"

    def _step_with_recovery(self, step_name: str, link: SerialLink,
                            fn: Callable[[], None],
                            context_msg: str) -> str:
        """Run `fn()` under disconnect-retry policy. Returns one of:
            'ok'        — step completed successfully
            'failed'    — step kept failing after max retries
            'different' — different device showed up; caller should abort
            'cancelled' — user cancelled; caller should abort
            'gone'      — device didn't come back; caller should abort
        Per spec, the step is idempotent — on every retry, fn() runs from
        the top (so e.g. config does cfg_getcfg → diff → batches → cfg_save
        all over again). Don't call this for firmware YMODEM — that has
        its own dedicated power-cycle recovery path.
        """
        attempts = 0
        while True:
            try:
                fn()
                return "ok"
            except DeviceDisconnected as e:
                self._logger.error(f"{step_name}: disconnected ({e})")
                attempts += 1
                if attempts > self.RECONNECT_RETRIES_PER_STEP:
                    self._has_error = True
                    self.errors.append(f"{step_name}: device disconnected, "
                                       f"reconnects exhausted")
                    return "failed"
                outcome = self._wait_for_reconnect(link, context_msg)
                if outcome == "cancelled": return "cancelled"
                if outcome == "different": return "different"
                if outcome == "timeout":
                    self._has_error = True
                    self.errors.append(f"{step_name}: device did not return")
                    return "gone"
                # outcome == "same" → loop and re-run fn from the top
                # Re-handshake (cfg_connect + auth) since fresh connection.
                try:
                    cmd_cfg_connect(link, timeout_s=1.5)
                    self._authenticate(link)
                except DeviceDisconnected:
                    continue  # loop will catch on next retry

    def _authenticate(self, link: SerialLink) -> bool:
        """Run sec_status → sec_login flow if device is secured. Updates
        self.opts.current_keyword to the keyword that actually unlocked the
        device. Returns True if authenticated (or unsecured), False if
        the user cancelled.

        Side effect — sets self._device_keyword to the device's CURRENT
        keyword value (working keyword on a secured device, empty string
        on an unsecured one). This is what the grid's Keyword column ends
        up showing if no later Set/Remove step overrides it."""
        sec = cmd_sec_status(link)
        if sec is None:
            raise RuntimeError("No <SECSTAT> response — not a Teltonika or wrong port")
        if not sec.has_keyword or sec.authenticated:
            # Unsecured (or somehow already authenticated). Either way the
            # device's keyword value to report is empty / whatever was used.
            if not sec.has_keyword:
                self._device_keyword = ""
            return True
        kw = self.opts.current_keyword.strip()
        attempt = 0
        while True:
            self._check_cancel_between_steps()
            if not kw:
                kw = self.ask_keyword(self.ident.imei) or ""
                if not kw: return False
            self.ui("Authentication", 10, f"Authenticating (attempt {attempt + 1})…")
            sec2 = cmd_sec_login(link, kw)
            if sec2 and sec2.authenticated:
                self.opts.current_keyword = kw
                # Device is secured WITH this keyword — that's the value
                # the grid will show if no Set/Remove step changes it.
                self._device_keyword = kw
                return True
            retries = sec2.retries if sec2 else "?"
            self.ui("Authentication", 10,
                    f"Wrong keyword (retries left: {retries}) — asking user…")
            kw = self.ask_keyword(self.ident.imei) or ""
            attempt += 1
            if not kw: return False

    def run(self) -> DeviceResult:
        res = DeviceResult(
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            port=self.port,
            imei=self.ident.imei,
            model=self.ident.model,
            firmware=self.ident.firmware,
        )

        # Single shared log file (always detailed). Workers append; the App
        # truncated the file on Start. The header banner makes it obvious
        # in the file where one device ends and the next begins.
        # Promote to an instance attribute so the disconnect-recovery code
        # paths can reach it without dragging a parameter everywhere.
        self._logger = Logger(self.opts.log_path or None, LOG_MODE_DETAILED)
        logger = self._logger   # short local alias keeps the rest readable
        logger.event("=" * 78)
        logger.event(f"Device IMEI={self.ident.imei} model={self.ident.model} "
                     f"port={self.port} firmware={self.ident.firmware}")
        logger.event("=" * 78)

        link = SerialLink(self.port, logger=logger)
        try:
            link.open()
        except Exception as e:
            self._has_error = True
            res.remarks = f"Cannot open port: {e}"
            logger.error(res.remarks); logger.close()
            res.has_error = self._has_error
            return res

        try:
            # ── Initial handshake + auth (retry-on-disconnect) ──────────
            def _initial_auth():
                self.ui("Authentication", 3, "Sending :cfg_connect…")
                cmd_cfg_connect(link, timeout_s=1.5)
                self.ui("Authentication", 5, "Reading security status…")
                if not self._authenticate(link):
                    raise CancelRequested()  # user cancelled the keyword prompt

            outcome = self._step_with_recovery(
                "Authentication", link, _initial_auth,
                context_msg="Device disconnected during authentication. "
                            "Re-plug the cable — we'll redo this step.")
            if outcome == "cancelled": raise CancelRequested()
            if outcome == "different":
                self._has_error = True
                res.remarks = "Device replaced during authentication."
                res.has_error = self._has_error
                return res
            if outcome == "gone":
                self._has_error = True
                res.remarks = "Device disconnected and did not return."
                res.has_error = self._has_error
                return res
            if outcome == "failed":
                self._has_error = True
                res.remarks = "Authentication failed after reconnects."
                res.has_error = self._has_error
                return res

            # ── Firmware ────────────────────────────────────────────────
            if self.opts.do_firmware and not self.skip_firmware_evt.is_set():
                self._check_cancel_between_steps()
                try:
                    self.ui("Firmware", 0,
                            f"Reading {Path(self.opts.firmware_path).name}…")
                    xim = Path(self.opts.firmware_path).read_bytes()
                    if xim[:4] != b"TLTF":
                        raise RuntimeError("Firmware file is not a .xim (TLTF) container")
                    self.ui("Firmware", 1, "Unwrapping .xim…")
                    container = xim_parse(xim)
                    self.ui("Firmware", 2,
                            f"Container unwrapped ({len(container['image']) // 1024} KB)")

                    # ── HOST-SIDE compatibility check ─────────────────
                    fw_text = (self.ident.firmware or "0.0.0").split(" ")[0]
                    ok, blockers = xim_check_compatibility(container, {
                        "imei": self.ident.imei,
                        "fw_version_text": fw_text,
                        "device_type": "FMB",
                    })
                    if not ok:
                        # Special signal to UI: open the dedicated "incompatible
                        # firmware" phase with a Skip-firmware option so config
                        # and keyword can still run. The UI sets either the
                        # skip flag or the cancel flag, then we re-check below.
                        logger.event("xim_check_compatibility refused: "
                                     + " | ".join(blockers))
                        msg = ("This firmware is not compatible with the connected "
                               "device:\n" + "\n".join("  • " + b for b in blockers))
                        self.ui("FwIncompatible", 0, msg)
                        # Wait for the user — either skip_firmware_evt is set
                        # (continue) or cancel_evt is set (skip rest).
                        while not (self.skip_firmware_evt.is_set() or
                                   self.cancel_evt.is_set()):
                            time.sleep(0.05)
                        if self.cancel_evt.is_set():
                            raise CancelRequested()
                        # Skip firmware, but record that we did.
                        self.errors.append("Firmware: skipped (incompatible)")
                    else:
                        self.ui("Firmware", 3, "Host-side compatibility check OK")
                        # ── Pre-flight + YMODEM transfer ──────────────
                        # The handshake's first 7 steps are reversible (device
                        # still in normal firmware) — disconnect there is
                        # treatable as any other step. Step 8+ + YMODEM are
                        # NOT reversible — disconnect there leaves the
                        # device in bootloader. We mark that case specially.
                        in_bootloader = [False]
                        def _flash_with_marker():
                            # When `.run_uart_boot_mode` (no `?`) is sent we
                            # consider ourselves in bootloader. We can't see
                            # the exact moment from out here, but the
                            # handshake hits that line BEFORE the data-packet
                            # loop, so we flip the marker right after the
                            # handshake returns successfully.
                            fw_handshake(link, container,
                                         lambda label: self.ui("Firmware", 4, label))
                            in_bootloader[0] = True
                            # Header + data packets + EOT + null (inner part
                            # of flash_firmware, sans handshake).
                            _fw_data_phase(
                                link, container,
                                Path(self.opts.firmware_path).name,
                                lambda pct, msg: self.ui("Firmware", pct, msg),
                                lambda: self.cancel_evt.is_set())
                        try:
                            _flash_with_marker()
                        except DeviceDisconnected as e:
                            if in_bootloader[0]:
                                # Mid-flash disconnect — bootloader limbo.
                                logger.error(f"Firmware mid-flash disconnect: {e}")
                                raise RuntimeError(
                                    "Device disconnected during firmware transfer. "
                                    "It is in bootloader mode and will auto-reboot "
                                    "to its old firmware after 5 minutes — or you "
                                    "can power-cycle it now (off for 5 s, on again) "
                                    "to recover faster.")
                            # Pre-bootloader disconnect — caller can retry.
                            raise
                        # ── 3-minute Flashing Phase countdown ─────────
                        self.ui("FlashingPhase", 0, "Flashing target firmware. Please wait…")
                        for elapsed in range(180):
                            if self.cancel_evt.is_set(): break
                            self.ui("FlashingPhase", elapsed, "")
                            time.sleep(1)
                        link.close()
                        # ── Wait for device to come back ──────────────
                        self.ui("FlashingPhase", 180, "Reconnecting to device after reboot…")
                        new_ident = None
                        deadline = time.monotonic() + 60
                        while time.monotonic() < deadline:
                            if self.cancel_evt.is_set(): raise CancelRequested()
                            new_ident = probe_port(self.port)
                            if new_ident: break
                            time.sleep(0.5)
                        if not new_ident:
                            raise RuntimeError(
                                "Device did not come back after the 3-minute "
                                "flash window. Power-cycle and re-run.")
                        # IMEI must still match — if a different device was
                        # plugged in during the wait, treat it as a new
                        # detection and abort this worker.
                        if new_ident.imei and new_ident.imei != self.ident.imei:
                            raise RuntimeError(
                                "A different device was plugged in during the "
                                "post-flash wait — aborting this device.")
                        link.open()
                        cmd_cfg_connect(link, timeout_s=1.5)
                        if not self._authenticate(link):
                            raise RuntimeError("Cancelled during post-flash re-auth.")
                        if container.get("version"):
                            v = container["version"]
                            res.firmware = (f"{v['major']:02d}.{v['minor']:02d}.{v['build']:02d}"
                                            + (f" Rev:{v['rev']}" if v.get("rev") else ""))
                        else:
                            res.firmware = new_ident.firmware or res.firmware
                except CancelRequested:
                    raise
                except DeviceDisconnected as e:
                    # Pre-bootloader disconnect (handshake hadn't returned).
                    # Try reconnect — if successful, surface as a soft "skip
                    # firmware" so the user can retry plug + continue with
                    # config/keyword. If it really doesn't come back, fail.
                    self._has_error = True
                    self.errors.append(f"Firmware: disconnected before bootloader ({e})")
                    logger.error(str(e))
                    outcome = self._wait_for_reconnect(
                        link,
                        "Device disconnected during firmware pre-flight. "
                        "Re-plug — we'll skip firmware and continue with "
                        "the remaining steps.")
                    if outcome == "cancelled": raise CancelRequested()
                    if outcome == "different":
                        self._has_error = True
                        res.remarks = "Device replaced during firmware pre-flight."
                        res.has_error = self._has_error
                        return res
                    if outcome == "gone":
                        self._has_error = True
                        res.remarks = "Device disconnected during firmware and did not return."
                        res.has_error = self._has_error
                        return res
                    # Got the same device back — skip firmware, continue.
                except Exception as e:
                    self._has_error = True
                    self.errors.append(f"Firmware: {e}")
                    logger.error(str(e))
                    self.ui("Firmware", -1, f"FAILED: {e}")
                    # UI lets the user Skip (continue with config+kw) or Cancel.
                    while not (self.skip_firmware_evt.is_set() or
                               self.cancel_evt.is_set()):
                        time.sleep(0.05)
                    if self.cancel_evt.is_set():
                        raise CancelRequested()

            # ── Config upload (with retry-on-disconnect) ────────────────
            if self.opts.do_config:
                self._check_cancel_between_steps()
                def _do_config():
                    self.ui("Config", 0, f"Reading {Path(self.opts.config_path).name}…")
                    raw = Path(self.opts.config_path).read_bytes()
                    cfg = parse_cfg_bytes(raw)
                    self.ui("Config", 2, f"Parsed {len(cfg.pids)} parameters from file")
                    saved, matched, modified, failed = upload_cfg(link, cfg,
                        lambda pct, msg: self.ui("Config", pct, msg))
                    res.config_file = Path(self.opts.config_path).name
                    # Friendly summary always shown in Remarks. Never sets
                    # has_error — partial device-side rejects are not errors
                    # (the device just declined some PIDs, that's its call).
                    if modified == 0:
                        self.errors.append(
                            f"Config: all {matched} params already match device")
                    else:
                        self.errors.append(
                            f"Config: {matched} params match already, "
                            f"{saved} of {modified} modified params written "
                            f"successfully")
                try:
                    outcome = self._step_with_recovery(
                        "Config", link, _do_config,
                        context_msg="Device disconnected during config upload. "
                                    "Re-plug — we'll redo the config from the start.")
                    if outcome == "cancelled": raise CancelRequested()
                    if outcome == "different":
                        self._has_error = True
                        res.remarks = "Device replaced during config upload."
                        return res
                    if outcome == "gone":
                        # _step_with_recovery already appended + set has_error.
                        pass
                except Exception as e:
                    if not isinstance(e, (CancelRequested,)):
                        self._has_error = True
                        self.errors.append(f"Config: {e}")
                        logger.error(str(e))
                        self.ui("Config", -1, f"FAILED: {e}")
                    else:
                        raise

            # ── Keyword action (with retry-on-disconnect) ───────────────
            if self.opts.keyword_action == "set":
                self._check_cancel_between_steps()
                def _do_set_keyword():
                    new_kw = self.opts.new_keyword.strip()
                    if not new_kw: raise RuntimeError("Empty new keyword")
                    self.ui("Keyword", 30, f"Setting keyword to '{new_kw}'…")
                    sec3 = cmd_sec_status(link)
                    if sec3 and sec3.has_keyword:
                        cur = self.opts.current_keyword.strip()
                        if not cur:
                            raise RuntimeError("Device is secured but no current keyword is known")
                        wire_cmd = "changekey"
                        sec4 = cmd_sec_changekey(link, new_kw, cur)
                    else:
                        wire_cmd = "setkey"
                        sec4 = cmd_sec_setkey(link, new_kw)
                    ok = (sec4 is not None
                          and sec4.is_valid_keyword == 1
                          and sec4.has_keyword == 1)
                    if not ok:
                        retries = sec4.retries if sec4 else "?"
                        raise RuntimeError(
                            f"Device rejected :sec_{wire_cmd} "
                            f"(result={sec4.is_valid_keyword if sec4 else '?'}, "
                            f"secured={sec4.has_keyword if sec4 else '?'}, "
                            f"retries left={retries})")
                    # Device's current keyword is now `new_kw`.
                    self._device_keyword = new_kw
                    self.ui("Keyword", 100, "Keyword set")
                try:
                    outcome = self._step_with_recovery(
                        "Keyword", link, _do_set_keyword,
                        context_msg="Device disconnected during keyword change. "
                                    "Re-plug — we'll redo the keyword step.")
                    if outcome == "cancelled": raise CancelRequested()
                    if outcome == "different":
                        self._has_error = True
                        res.remarks = "Device replaced during keyword step."
                        res.has_error = self._has_error
                        return res
                except Exception as e:
                    if not isinstance(e, CancelRequested):
                        self._has_error = True
                        self.errors.append(f"Keyword set: {e}")
                        logger.error(str(e))
                        self.ui("Keyword", -1, f"FAILED: {e}")
                    else:
                        raise

            elif self.opts.keyword_action == "remove":
                self._check_cancel_between_steps()
                def _do_remove_keyword():
                    # Per spec: don't blindly issue :sec_removekey. If the
                    # device has no keyword to begin with, skip silently —
                    # nothing to do. Only secured devices get the actual
                    # remove command.
                    self.ui("Keyword", 20, "Checking security status…")
                    sec_now = cmd_sec_status(link)
                    if sec_now is not None and sec_now.has_keyword == 0:
                        link.logger.event(
                            "Keyword remove: device has no keyword — skipping")
                        # No-op — device's current keyword is (still) empty.
                        self._device_keyword = ""
                        self.ui("Keyword", 100, "No keyword on device — skipped")
                        return
                    self.ui("Keyword", 50, "Removing keyword…")
                    sec5 = cmd_sec_removekey(link)
                    ok = (sec5 is not None
                          and sec5.is_valid_keyword == 1
                          and sec5.has_keyword == 0)
                    if not ok:
                        raise RuntimeError(
                            f"Device did not accept :sec_removekey "
                            f"(result={sec5.is_valid_keyword if sec5 else '?'}, "
                            f"secured={sec5.has_keyword if sec5 else '?'})")
                    # Device now has no keyword.
                    self._device_keyword = ""
                    self.ui("Keyword", 100, "Keyword removed")
                try:
                    outcome = self._step_with_recovery(
                        "Keyword", link, _do_remove_keyword,
                        context_msg="Device disconnected during keyword removal. "
                                    "Re-plug — we'll redo the keyword step.")
                    if outcome == "cancelled": raise CancelRequested()
                    if outcome == "different":
                        self._has_error = True
                        res.remarks = "Device replaced during keyword step."
                        res.has_error = self._has_error
                        return res
                except Exception as e:
                    if not isinstance(e, CancelRequested):
                        self._has_error = True
                        self.errors.append(f"Keyword remove: {e}")
                        logger.error(str(e))
                        self.ui("Keyword", -1, f"FAILED: {e}")
                    else:
                        raise

            # Keyword column reflects the device's CURRENT keyword value
            # (post-batch). Empty string means the device has no keyword.
            # This is a data field, not a status — it's consumed by
            # downstream CSV-import platforms.
            res.keyword = self._device_keyword
            if self.errors:
                res.remarks = "; ".join(self.errors)
            res.has_error = self._has_error
            logger.event(f"Device done: keyword='{res.keyword}'  "
                         f"has_error={res.has_error}  remarks={res.remarks}")
            return res
        except CancelRequested:
            self._has_error = True
            res.remarks = "Cancelled by user"
            res.has_error = self._has_error
            return res
        finally:
            link.close()
            logger.close()


# ─────────────────────────────────────────────────────────────────────────────
# Tkinter UI
# ─────────────────────────────────────────────────────────────────────────────

GRID_COLS = ("ts", "imei", "model", "firmware", "config", "keyword", "remarks")
GRID_HEAD = ("Time", "IMEI", "Model", "Firmware", "Config file", "Keyword", "Remarks")
GRID_W    = (140, 130, 80, 110, 180, 90, 360)


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("GLOBO360 Teltonika Batch Configurator")
        root.geometry("1180x720")  # used if maximize fails (e.g. on a window manager that doesn't support `zoomed`)
        root.minsize(900, 560)
        # Start maximized. Windows / Mac use `zoomed`; Linux/Tk uses the
        # `-zoomed` attribute. Try both and don't die on whichever rejects it.
        try:
            root.state("zoomed")
        except tk.TclError:
            try:
                root.attributes("-zoomed", True)
            except tk.TclError:
                pass

        self.results: list[DeviceResult] = []
        self._build_form()
        self._build_grid()

        # Cross-thread message bus: worker pushes (kind, payload), UI consumes.
        self._msg_q: queue.Queue = queue.Queue()
        self.root.after(50, self._drain_msgs)

        # Single morphing modal — replaces the older detect/progress/duplicate/
        # ask_keyword popups. Phase is one of: waiting | duplicate | grace |
        # step | flashing | ask_keyword | fw_fail.
        self._modal: Optional[tk.Toplevel] = None
        self._modal_body: Optional[ttk.Frame] = None
        self._modal_phase: Optional[str] = None
        self._modal_W, self._modal_H = 540, 360
        self._cur_ident: Optional[DeviceIdentity] = None
        self._cur_port: Optional[str] = None

        self._batch_active = False
        self._last_port: Optional[str] = None
        self._last_port_set: Optional[set] = None
        self._cancel_evt = threading.Event()
        self._skip_firmware_evt = threading.Event()
        self._worker_thread: Optional[threading.Thread] = None
        self._detect_thread: Optional[threading.Thread] = None

    # ── Form (top half) ──────────────────────────────────────────────────
    def _build_form(self):
        bar = ttk.Frame(self.root, padding=10)
        bar.pack(fill="x")

        # Firmware row
        # Both file-picker rows share the same 3-column grid so the Browse
        # buttons and path entries line up vertically across the two rows.
        files_grid = ttk.Frame(bar); files_grid.pack(fill="x", pady=3)
        files_grid.columnconfigure(0, weight=0, minsize=180)  # checkbox col, fixed
        files_grid.columnconfigure(1, weight=0)               # browse btn, fixed
        files_grid.columnconfigure(2, weight=1)               # path entry, stretches

        self.do_fw_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(files_grid, text="Firmware Update (.xim)",
                        variable=self.do_fw_var,
                        command=self._on_form_change
                        ).grid(row=0, column=0, sticky="w", pady=3)
        self.fw_path_var = tk.StringVar(value="")
        self.fw_browse_btn = ttk.Button(
            files_grid, text="Browse…",
            # When the user picks a firmware file, auto-tick the firmware
            # checkbox — they obviously want that step enabled.
            command=lambda: self._browse(self.fw_path_var,
                                         [("Teltonika firmware", "*.xim"),
                                          ("All files", "*.*")],
                                         check_var=self.do_fw_var))
        self.fw_browse_btn.grid(row=0, column=1, padx=6, pady=3)
        self.fw_entry = ttk.Entry(files_grid, textvariable=self.fw_path_var,
                                  state="readonly")
        self.fw_entry.grid(row=0, column=2, sticky="ew", padx=4, pady=3)

        self.do_cfg_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(files_grid, text="Config Upload (.cfg)",
                        variable=self.do_cfg_var,
                        command=self._on_form_change
                        ).grid(row=1, column=0, sticky="w", pady=3)
        self.cfg_path_var = tk.StringVar(value="")
        self.cfg_browse_btn = ttk.Button(
            files_grid, text="Browse…",
            # Picking a config auto-ticks the config checkbox.
            command=lambda: self._browse(self.cfg_path_var,
                                         [("Teltonika config", "*.cfg"),
                                          ("All files", "*.*")],
                                         check_var=self.do_cfg_var))
        self.cfg_browse_btn.grid(row=1, column=1, padx=6, pady=3)
        self.cfg_entry = ttk.Entry(files_grid, textvariable=self.cfg_path_var,
                                   state="readonly")
        self.cfg_entry.grid(row=1, column=2, sticky="ew", padx=4, pady=3)

        # Group row — single text tag applied to every device processed in
        # this batch. Lands in the XLSX export's Group column. No checkbox:
        # whatever's in the textbox at Start time gets stamped on every
        # device. Empty textbox → empty Group column. Lines up under the
        # file rows by reusing the same grid columns.
        ttk.Label(files_grid, text="Group Name",
                  anchor="w").grid(row=2, column=0, sticky="w", pady=3, padx=(22, 0))
        self.group_name_var = tk.StringVar(value="")
        self.group_name_entry = ttk.Entry(files_grid,
                                          textvariable=self.group_name_var)
        self.group_name_entry.grid(row=2, column=1, columnspan=2,
                                   sticky="ew", padx=(6, 4), pady=3)

        # Routing block — Server IP + Port applied at export time to every
        # row in the XLSX. No checkbox: whatever's in these fields at the
        # moment Export is clicked gets stamped onto the exported rows.
        # Empty is fine — the columns just come out blank.
        routing_outer = ttk.LabelFrame(bar, text="Routing", padding=8)
        routing_outer.pack(fill="x", pady=(6, 0))

        ip_row = ttk.Frame(routing_outer); ip_row.pack(fill="x", pady=2)
        ttk.Label(ip_row, text="Routing IP:", width=18,
                  anchor="w").pack(side="left")
        self.routing_ip_var = tk.StringVar(value="")
        self.routing_ip_entry = ttk.Entry(ip_row,
                                          textvariable=self.routing_ip_var,
                                          width=40)
        self.routing_ip_entry.pack(side="left", fill="x", expand=True)

        port_row = ttk.Frame(routing_outer); port_row.pack(fill="x", pady=2)
        ttk.Label(port_row, text="Routing Port:", width=18,
                  anchor="w").pack(side="left")
        self.routing_port_var = tk.StringVar(value="")
        self.routing_port_entry = ttk.Entry(port_row,
                                            textvariable=self.routing_port_var,
                                            width=12)
        self.routing_port_entry.pack(side="left")

        # Password block.
        # Current Keyword sits on its OWN row at the top — used to authenticate
        # any secured device before firmware/config, and required for Remove,
        # and (when the device is already secured) re-used as the "current"
        # half of an internal :sec_changekey for Set.
        pw_outer = ttk.LabelFrame(bar, text="Password", padding=8)
        pw_outer.pack(fill="x", pady=6)

        # Current Keyword (always editable, sits above the action choices)
        cur_row = ttk.Frame(pw_outer); cur_row.pack(fill="x", pady=2)
        ttk.Label(cur_row, text="Current Keyword:", width=18, anchor="w").pack(side="left")
        self.cur_kw_var = tk.StringVar(value="")
        self.cur_kw_entry = ttk.Entry(cur_row, textvariable=self.cur_kw_var, width=24)
        self.cur_kw_entry.pack(side="left")
        ttk.Label(cur_row, text="(leave blank if the device has no keyword yet)",
                  foreground="#666").pack(side="left", padx=8)

        # Radio choices for what to do with the keyword
        self.kw_action_var = tk.StringVar(value="none")
        set_row = ttk.Frame(pw_outer); set_row.pack(fill="x", pady=1)
        ttk.Radiobutton(set_row, text="Set Keyword", variable=self.kw_action_var, value="set",
                        command=self._on_form_change).pack(side="left")
        ttk.Label(set_row, text="New Keyword:").pack(side="left", padx=(20, 4))
        self.new_kw_var = tk.StringVar(value="")
        self.new_kw_entry = ttk.Entry(set_row, textvariable=self.new_kw_var, width=24)
        self.new_kw_entry.pack(side="left")

        rem_row = ttk.Frame(pw_outer); rem_row.pack(fill="x", pady=1)
        ttk.Radiobutton(rem_row, text="Remove Keyword",
                        variable=self.kw_action_var, value="remove",
                        command=self._on_form_change).pack(side="left")

        none_row = ttk.Frame(pw_outer); none_row.pack(fill="x", pady=1)
        ttk.Radiobutton(none_row, text="No changes", variable=self.kw_action_var, value="none",
                        command=self._on_form_change).pack(side="left")

        # Action bar
        act_row = ttk.Frame(bar); act_row.pack(fill="x", pady=(10, 0))
        self.start_btn = ttk.Button(act_row, text="Start", command=self._start_batch, width=14)
        self.start_btn.pack(side="left")
        ttk.Button(act_row, text="Export XLSX",
                   command=self._save_xlsx, width=16).pack(side="left", padx=8)
        ttk.Button(act_row, text="Reset",
                   command=self._reset_all, width=12).pack(side="left")
        self.status_lbl = ttk.Label(act_row, text="Ready.", foreground="#444")
        self.status_lbl.pack(side="left", padx=12)

        self._on_form_change()

    def _on_form_change(self):
        # New Keyword only matters when the action is Set.
        st_set = "normal" if self.kw_action_var.get() == "set" else "disabled"
        self.new_kw_entry.configure(state=st_set)
        # Current Keyword is ALWAYS editable — three roles:
        #   (a) authenticate a secured device before firmware / config
        #   (b) the "current" half of a Set on an already-secured device
        #   (c) the auth before Remove
        self.cur_kw_entry.configure(state="normal")

    def _browse(self, var: tk.StringVar, types,
                check_var: Optional[tk.BooleanVar] = None):
        """Open a file picker. If a file is selected and a `check_var` is
        provided, tick the matching checkbox automatically — picking a file
        is an implicit "yes, I want this step enabled"."""
        path = filedialog.askopenfilename(filetypes=types)
        if not path: return
        var.set(path)
        if check_var is not None:
            check_var.set(True)
            self._on_form_change()

    # ── Grid (bottom half) ───────────────────────────────────────────────
    def _build_grid(self):
        frm = ttk.LabelFrame(self.root, text="Configured devices", padding=6)
        frm.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree = ttk.Treeview(frm, columns=GRID_COLS, show="headings",
                                 height=18, selectmode="browse")
        for i, (c, h, w) in enumerate(zip(GRID_COLS, GRID_HEAD, GRID_W)):
            # Header anchor "w" matches the row-data anchor below so the
            # column label sits over the start of the values, not centered.
            self.tree.heading(c, text=h, anchor="w")
            self.tree.column(c, width=w, anchor="w")
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm.rowconfigure(0, weight=1); frm.columnconfigure(0, weight=1)
        # Tag for error rows
        self.tree.tag_configure("error", background="#fee4e4")
        self.tree.tag_configure("ok",    background="#e8f7e8")

    def _append_result(self, r: DeviceResult):
        self.results.append(r)
        # Red row only for GENUINE failures (auth/disconnect/firmware). Config
        # partial writes (some PIDs rejected) and other informational remarks
        # keep the row green — the device did its job, it just chose to reject
        # some params, which isn't our problem.
        tag = "error" if r.has_error else "ok"
        # Fall back to "NO ERRORS" placeholder for an empty remarks cell so
        # the operator can see at a glance that nothing was worth logging.
        remark_cell = r.remarks if r.remarks else "NO ERRORS"
        self.tree.insert("", "end",
                         values=(r.timestamp, r.imei, r.model, r.firmware,
                                 r.config_file, r.keyword, remark_cell),
                         tags=(tag,))

    # ── Validation ───────────────────────────────────────────────────────
    def _validate(self) -> Optional[str]:
        if self.do_fw_var.get():
            p = self.fw_path_var.get().strip()
            if not p or not Path(p).is_file():
                return "Please pick a firmware (.xim) file."
            if not p.lower().endswith(".xim"):
                return "Firmware must be a .xim file."
        if self.do_cfg_var.get():
            p = self.cfg_path_var.get().strip()
            if not p or not Path(p).is_file():
                return "Please pick a config (.cfg) file."
        if self.kw_action_var.get() == "set":
            if not self.new_kw_var.get().strip():
                return "Set Keyword requires a new keyword."
        # NOTE: we deliberately do NOT require a Current Keyword for the
        # Remove action up-front. Per-device behaviour decides at runtime:
        # if the device is already unsecured, the Remove step skips itself
        # silently; if it IS secured and no current was provided, the auth
        # step pops the wrong-keyword modal so the operator can supply one
        # just for that device.
        if not (self.do_fw_var.get() or self.do_cfg_var.get()
                or self.kw_action_var.get() != "none"):
            return "Pick at least one action (firmware, config, or keyword change)."
        return None

    def _opts_snapshot(self) -> BatchOptions:
        # Single shared log file next to the running app (source file in
        # dev, .exe in packaged builds — see _app_dir). Always full detail.
        # The app truncates this file on Start; each worker opens it in
        # append.
        log_path = str(_app_dir() / "batch_configurator.log")
        return BatchOptions(
            do_firmware=self.do_fw_var.get(),
            firmware_path=self.fw_path_var.get().strip(),
            do_config=self.do_cfg_var.get(),
            config_path=self.cfg_path_var.get().strip(),
            keyword_action=self.kw_action_var.get(),
            new_keyword=self.new_kw_var.get().strip(),
            current_keyword=self.cur_kw_var.get().strip(),
            log_path=log_path,
        )

    # ── Batch start / loop ───────────────────────────────────────────────
    def _start_batch(self):
        err = self._validate()
        if err:
            messagebox.showerror("Cannot start", err); return
        # Truncate the shared log file. Workers open it in append; this
        # makes Start = a fresh session, no historical clutter.
        # Logging is best-effort — disk full / file locked / permissions /
        # whatever: silently skip. The batch is the important thing.
        log_path = _app_dir() / "batch_configurator.log"
        try:
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"==== batch started {datetime.now().isoformat()} ====\n")
        except Exception:
            pass
        self._batch_active = True
        self._lock_form(True)
        self._open_modal()
        self._set_phase("waiting")
        self._begin_detect(use_last_port=False)

    def _lock_form(self, locked: bool):
        st = "disabled" if locked else "normal"
        for w in (self.start_btn, self.fw_browse_btn, self.cfg_browse_btn):
            w.configure(state=st)

    # ── Single morphing modal ────────────────────────────────────────────
    def _open_modal(self):
        if self._modal: return
        m = tk.Toplevel(self.root)
        m.title("GLOBO360 Teltonika Batch Configurator")
        W, H = self._modal_W, self._modal_H
        self.root.update_idletasks()
        rx, ry = self.root.winfo_rootx(), self.root.winfo_rooty()
        rw, rh = self.root.winfo_width(), self.root.winfo_height()
        x = rx + max(0, (rw - W) // 2)
        y = ry + max(0, (rh - H) // 2)
        m.geometry(f"{W}x{H}+{x}+{y}")
        m.transient(self.root); m.grab_set()
        m.protocol("WM_DELETE_WINDOW", lambda: None)
        self._modal = m
        self._modal_body = ttk.Frame(m, padding=20)
        self._modal_body.pack(fill="both", expand=True)
        self._modal_phase = None

    def _close_modal(self):
        if self._modal:
            try: self._modal.grab_release()
            except Exception: pass
            self._modal.destroy()
            self._modal = None
            self._modal_body = None
            self._modal_phase = None

    def _clear_modal_body(self):
        if not self._modal_body: return
        for w in list(self._modal_body.winfo_children()):
            try: w.destroy()
            except Exception: pass
        # Strip any phase-specific key bindings so they don't leak.
        try:
            self._modal.unbind("<Return>"); self._modal.unbind("<Escape>")
        except Exception:
            pass

    def _set_phase(self, phase: str, **kw):
        """Switch the modal to a new phase by re-rendering its body."""
        if not self._modal: self._open_modal()
        self._modal_phase = phase
        self._clear_modal_body()
        renderer = getattr(self, f"_render_{phase}", None)
        if renderer:
            renderer(**kw)

    # ── Phase: waiting ──────────────────────────────────────────────────
    def _render_waiting(self, headline: str = "Plug in a Teltonika device…",
                        sub: str = "Watching USB for changes…"):
        self._modal.title("GLOBO360 Teltonika Batch Configurator — Waiting")
        ttk.Label(self._modal_body, text=headline,
                  font=("Segoe UI", 12)).pack(pady=(14, 4))
        self._waiting_sub = ttk.Label(self._modal_body, text=sub, foreground="#555")
        self._waiting_sub.pack(pady=2)
        bar = ttk.Progressbar(self._modal_body, mode="indeterminate", length=380)
        bar.pack(pady=24); bar.start(10)
        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(10, 4))
        ttk.Button(bot, text="Close", width=14,
                   command=self._end_batch).pack(side="right", padx=4)

    def _update_waiting_sub(self, text: str):
        if self._modal_phase == "waiting" and hasattr(self, "_waiting_sub"):
            try: self._waiting_sub.config(text=text)
            except Exception: pass

    # ── Phase: duplicate (zero interaction — keeps detect running) ───────
    def _render_duplicate(self, ident: DeviceIdentity,
                          prev: Optional[DeviceResult]):
        self._modal.title("GLOBO360 Teltonika Batch Configurator — Already configured")
        ok_text = ("with no errors" if (prev and not prev.has_error)
                   else "with errors")
        ttk.Label(self._modal_body, text="Device already configured",
                  font=("Segoe UI", 16, "bold"),
                  foreground="#a06400").pack(pady=(20, 8))
        ttk.Label(self._modal_body,
                  text=f"IMEI {ident.imei}  ({ident.model or 'device'})",
                  foreground="#444", font=("Segoe UI", 10)).pack()
        if prev:
            ttk.Label(self._modal_body,
                      text=f"Configured at {prev.timestamp} {ok_text}",
                      foreground="#666").pack(pady=(2, 16))
        ttk.Separator(self._modal_body, orient="horizontal"
                      ).pack(fill="x", padx=20, pady=4)
        ttk.Label(self._modal_body,
                  text="Plug in a different device to continue.",
                  foreground="#555").pack(pady=(10, 4))
        bar = ttk.Progressbar(self._modal_body, mode="indeterminate", length=320)
        bar.pack(pady=10); bar.start(10)
        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(10, 4))
        ttk.Button(bot, text="Close", width=14,
                   command=self._end_batch).pack(side="right", padx=4)

    # ── Phase: grace (5s countdown before processing starts) ────────────
    def _render_grace(self, ident: DeviceIdentity, port: str, seconds_left: int):
        self._modal.title(f"GLOBO360 Teltonika Batch Configurator — Starting on {port}")
        ttk.Label(self._modal_body,
                  text=f"Detected {ident.model or 'device'}",
                  font=("Segoe UI", 12, "bold")).pack(pady=(14, 2))
        ttk.Label(self._modal_body,
                  text=f"IMEI {ident.imei}   on {port}",
                  foreground="#555").pack()
        self._grace_label = ttk.Label(
            self._modal_body,
            text=self._grace_text(seconds_left),
            font=("Segoe UI", 11), foreground="#444")
        self._grace_label.pack(pady=22)
        bar = ttk.Progressbar(self._modal_body, mode="determinate",
                              length=360, maximum=PRE_START_DELAY_S)
        bar["value"] = PRE_START_DELAY_S - seconds_left
        bar.pack()
        self._grace_bar = bar
        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(10, 4))
        ttk.Button(bot, text="Cancel", width=14,
                   command=self._end_batch).pack(side="right", padx=4)

    def _grace_text(self, n: int) -> str:
        return (f"Starting in {n} second{'s' if n != 1 else ''} — "
                "keep the cable still…")

    def _update_grace(self, seconds_left: int):
        if self._modal_phase == "grace" and hasattr(self, "_grace_label"):
            self._grace_label.config(text=self._grace_text(seconds_left))
            self._grace_bar["value"] = PRE_START_DELAY_S - seconds_left

    # ── Phase: step (current operation progress) ────────────────────────
    def _render_step(self, ident: DeviceIdentity, port: str,
                     step: str, pct: int, detail: str):
        self._modal.title(f"GLOBO360 Teltonika Batch Configurator — {step}")
        self._cur_ident, self._cur_port = ident, port
        ttk.Label(
            self._modal_body,
            text=f"{ident.model or '(unknown)'}    IMEI {ident.imei}    on {port}",
            font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self._step_lbl = ttk.Label(self._modal_body, text=f"{step}  ({pct}%)",
                                   font=("Segoe UI", 11))
        self._step_lbl.pack(pady=(12, 4), anchor="w")
        self._detail_lbl = ttk.Label(self._modal_body, text=detail, foreground="#555",
                                     wraplength=self._modal_W - 60, justify="left")
        self._detail_lbl.pack(pady=(0, 8), anchor="w")
        self._pbar = ttk.Progressbar(self._modal_body, mode="determinate",
                                     length=self._modal_W - 80, maximum=100)
        self._pbar["value"] = pct
        self._pbar.pack()
        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(14, 4))
        ttk.Button(bot, text="Cancel", width=14,
                   command=self._on_cancel_clicked).pack(side="right", padx=4)

    def _update_step(self, step: str, pct: int, detail: str):
        if self._modal_phase != "step":
            if self._cur_ident and self._cur_port:
                self._set_phase("step", ident=self._cur_ident, port=self._cur_port,
                                step=step, pct=pct, detail=detail)
            return
        self._modal.title(f"GLOBO360 Teltonika Batch Configurator — {step}")
        self._step_lbl.config(text=f"{step}  ({pct}%)")
        self._detail_lbl.config(text=detail, foreground="#555")
        self._pbar["value"] = pct

    # ── Phase: flashing (3-min countdown) ───────────────────────────────
    def _render_flashing(self):
        self._modal.title("GLOBO360 Teltonika Batch Configurator — Flashing firmware")
        ttk.Label(self._modal_body, text="Flashing target firmware.",
                  font=("Segoe UI", 14, "bold")).pack(pady=(18, 2))
        ttk.Label(self._modal_body, text="Please wait…",
                  foreground="#777").pack(pady=(0, 22))
        self._cd_label = ttk.Label(self._modal_body, text="03:00",
                                   font=("Consolas", 36, "bold"))
        self._cd_label.pack(pady=(0, 22))
        ttk.Label(self._modal_body,
                  text="DO NOT CLOSE THIS WINDOW UNTIL COMPLETE",
                  foreground="#ef4444", font=("Segoe UI", 10, "bold")).pack()

    def _update_flashing(self, seconds_left: int):
        if self._modal_phase == "flashing" and hasattr(self, "_cd_label"):
            mm = max(0, seconds_left) // 60
            ss = max(0, seconds_left) % 60
            self._cd_label.config(text=f"{mm:02d}:{ss:02d}")

    # ── Phase: ask_keyword (worker waits on the threading.Event) ────────
    def _render_ask_keyword(self, imei: str, result_box: dict,
                            evt: threading.Event):
        self._modal.title("GLOBO360 Teltonika Batch Configurator — Wrong keyword")
        ttk.Label(self._modal_body, text="Wrong current keyword",
                  font=("Segoe UI", 12, "bold"),
                  foreground="#c00").pack(anchor="w", pady=(4, 2))
        ttk.Label(self._modal_body, text=f"IMEI {imei}",
                  foreground="#555").pack(anchor="w")
        ttk.Label(self._modal_body,
                  text="Enter the correct current keyword for this device, "
                       "or cancel to skip it (the batch continues with the next).",
                  foreground="#444", wraplength=self._modal_W - 40,
                  justify="left").pack(anchor="w", pady=(10, 6))
        kw_var = tk.StringVar(value="")
        ent = ttk.Entry(self._modal_body, textvariable=kw_var, width=30)
        ent.pack(anchor="w"); ent.focus_set()

        def _submit():
            result_box["value"] = kw_var.get().strip()
            evt.set()
        def _cancel():
            result_box["value"] = None
            evt.set()

        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(20, 4))
        ttk.Button(bot, text="Cancel this device", width=20,
                   command=_cancel).pack(side="right", padx=4)
        ttk.Button(bot, text="Submit", width=12,
                   command=_submit).pack(side="right", padx=4)
        self._modal.bind("<Return>", lambda e: _submit())
        self._modal.bind("<Escape>", lambda e: _cancel())

    # ── Phase: firmware failed (skip or cancel) ─────────────────────────
    def _render_fw_fail(self, detail: str):
        self._modal.title("GLOBO360 Teltonika Batch Configurator — Firmware failed")
        ttk.Label(self._modal_body, text="Firmware update failed",
                  font=("Segoe UI", 12, "bold"),
                  foreground="#c00").pack(anchor="w")
        ttk.Label(self._modal_body, text=detail, foreground="#444",
                  wraplength=self._modal_W - 40,
                  justify="left").pack(anchor="w", pady=8)
        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(20, 4))
        ttk.Button(bot, text="Cancel this device", width=20,
                   command=lambda: self._cancel_evt.set()
                   ).pack(side="right", padx=4)
        ttk.Button(bot, text="Skip firmware", width=14,
                   command=lambda: self._skip_firmware_evt.set()
                   ).pack(side="right", padx=4)

    # ── Phase: reconnecting (device disconnected mid-step) ───────────────
    def _render_reconnecting(self, detail: str, secs_left: int):
        self._modal.title("GLOBO360 Teltonika Batch Configurator — Device disconnected")
        ttk.Label(self._modal_body, text="Device disconnected",
                  font=("Segoe UI", 14, "bold"),
                  foreground="#c00").pack(pady=(14, 6))
        self._recon_msg = ttk.Label(self._modal_body, text=detail,
                                    foreground="#444",
                                    wraplength=self._modal_W - 40,
                                    justify="left")
        self._recon_msg.pack(pady=(0, 12))
        self._recon_secs = ttk.Label(self._modal_body,
                                     text=f"Waiting {secs_left} s for device "
                                          "to come back…",
                                     foreground="#555")
        self._recon_secs.pack(pady=(2, 8))
        bar = ttk.Progressbar(self._modal_body, mode="indeterminate", length=320)
        bar.pack(pady=4); bar.start(10)
        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(16, 4))
        ttk.Button(bot, text="Cancel this device", width=20,
                   command=lambda: self._cancel_evt.set()
                   ).pack(side="right", padx=4)

    def _update_reconnecting(self, secs_left: int, detail: str):
        if hasattr(self, "_recon_secs"):
            try:
                self._recon_secs.config(
                    text=f"Waiting {secs_left} s for device to come back…")
                self._recon_msg.config(text=detail)
            except Exception: pass

    # ── Phase: firmware not compatible (pre-flash) ──────────────────────
    def _render_fw_compat_fail(self, detail: str):
        self._modal.title("GLOBO360 Teltonika Batch Configurator — Firmware not compatible")
        ttk.Label(self._modal_body,
                  text="Device not compatible with this firmware",
                  font=("Segoe UI", 12, "bold"),
                  foreground="#a06400").pack(anchor="w", pady=(2, 8))
        ttk.Label(self._modal_body, text=detail, foreground="#444",
                  wraplength=self._modal_W - 40,
                  justify="left").pack(anchor="w")
        ttk.Label(self._modal_body,
                  text="\nYou can skip the firmware and still apply the config "
                       "and keyword changes to this device, or cancel and try a "
                       "different one.",
                  foreground="#666",
                  wraplength=self._modal_W - 40,
                  justify="left").pack(anchor="w", pady=(8, 0))
        bot = ttk.Frame(self._modal_body); bot.pack(side="bottom", fill="x", pady=(20, 4))
        ttk.Button(bot, text="Cancel this device", width=20,
                   command=lambda: self._cancel_evt.set()
                   ).pack(side="right", padx=4)
        ttk.Button(bot, text="Skip firmware, continue", width=22,
                   command=lambda: self._skip_firmware_evt.set()
                   ).pack(side="right", padx=4)

    # ── Detection plumbing — pushes messages into the queue, which
    # `_drain_msgs` translates into modal phase changes. ──────────────────
    def _begin_detect(self, use_last_port: bool):
        if not self._batch_active: return
        self._detect_thread = threading.Thread(
            target=self._detect_worker, args=(use_last_port,), daemon=True)
        self._detect_thread.start()

    def _detect_worker(self, use_last_port: bool):
        # 1) If we know the last port and a port-set change just happened,
        #    try that port first for LAST_PORT_WAIT_S seconds.
        if use_last_port and self._last_port:
            self._msg_q.put(("detect_sub", f"New USB change — retrying {self._last_port}…"))
            deadline = time.monotonic() + LAST_PORT_WAIT_S
            while time.monotonic() < deadline and self._batch_active:
                if self._last_port in all_current_ports():
                    ident = probe_port(self._last_port)
                    if ident:
                        self._msg_q.put(("detected", (self._last_port, ident)))
                        return
                time.sleep(0.4)

        # 2) Watch for port-collection changes. When ports change, retry the
        #    last known port first; otherwise full Teltonika scan.
        prev = all_current_ports()
        self._msg_q.put(("detect_sub", "Watching USB for changes…"))
        while self._batch_active:
            time.sleep(PORT_POLL_MS / 1000.0)
            cur = all_current_ports()
            if cur == prev:
                continue
            added = cur - prev
            prev = cur
            if not added:
                continue
            # New port(s) appeared. Quick probe each that matches Teltonika VID/PID.
            self._msg_q.put(("detect_sub", f"New USB port(s) detected ({len(added)}). Probing…"))
            # Give Windows a beat to fully enumerate.
            time.sleep(0.5)
            # Prefer the last known port if it's back; otherwise scan all Teltonika ports.
            if self._last_port and self._last_port in cur:
                ident = probe_port(self._last_port)
                if ident:
                    self._msg_q.put(("detected", (self._last_port, ident)))
                    return
            found = find_config_port()
            if found:
                self._msg_q.put(("detected", found))
                return

    # ── Run one device end-to-end ────────────────────────────────────────
    def _on_device_detected(self, port_and_ident):
        port, ident = port_and_ident
        self._last_port = port
        self._last_port_set = all_current_ports()

        # Duplicate IMEI? Switch to the zero-interaction "duplicate" phase,
        # keep the detect thread running so the next (different) device
        # auto-advances to grace without any clicks.
        if ident.imei and any(r.imei == ident.imei for r in self.results):
            prev = next((r for r in self.results if r.imei == ident.imei), None)
            self._set_phase("duplicate", ident=ident, prev=prev)
            self._begin_detect(use_last_port=False)
            return

        # Otherwise → 5-second grace then processing.
        self._grace_left = PRE_START_DELAY_S
        self._set_phase("grace", ident=ident, port=port,
                        seconds_left=self._grace_left)
        self.root.after(1000, self._grace_tick)

    def _grace_tick(self):
        if not self._batch_active: return
        if self._modal_phase != "grace": return  # user cancelled or phase changed
        self._grace_left -= 1
        if self._grace_left <= 0:
            self._begin_processing()
            return
        self._update_grace(self._grace_left)
        self.root.after(1000, self._grace_tick)

    def _begin_processing(self):
        port = self._last_port
        ident = probe_port(port)
        if not ident:
            # Device disconnected during grace — back to waiting silently.
            self._set_phase("waiting",
                            headline="Device disconnected during the grace period.",
                            sub="Plug in a Teltonika device…")
            self._begin_detect(use_last_port=True)
            return
        self._cancel_evt.clear()
        self._skip_firmware_evt.clear()
        self._set_phase("step", ident=ident, port=port,
                        step="Authentication", pct=0, detail="Starting…")
        opts = self._opts_snapshot()
        worker = DeviceWorker(port, ident, opts,
                              self._ui_progress, self._cancel_evt,
                              self._skip_firmware_evt, self._ask_keyword)
        self._worker_thread = threading.Thread(
            target=self._run_worker, args=(worker,), daemon=True)
        self._worker_thread.start()

    def _run_worker(self, worker: DeviceWorker):
        try:
            result = worker.run()
        except Exception as e:
            result = DeviceResult(
                timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                port=worker.port, imei=worker.ident.imei,
                model=worker.ident.model, firmware=worker.ident.firmware,
                config_file="", keyword="",
                remarks=f"Unexpected error: {e}",
                has_error=True,
            )
        self._msg_q.put(("device_done", result))

    def _ui_progress(self, step: str, pct: int, detail: str):
        """Called from worker thread."""
        self._msg_q.put(("progress", (step, pct, detail)))

    def _ask_keyword(self, imei: str) -> Optional[str]:
        """Called from the WORKER thread when auth fails. Blocks until the UI
        thread shows the modal and the user submits or cancels. Returns the
        new keyword to try, or None to skip this device."""
        evt = threading.Event()
        result_box: dict = {"value": None}
        self._msg_q.put(("ask_keyword", (imei, result_box, evt)))
        evt.wait()
        return result_box["value"]

    def _on_cancel_clicked(self):
        if messagebox.askyesno(
            "Cancel this device?",
            "The current step will finish first. After it does, this device "
            "will be cancelled.\n\nContinue?",
            parent=self._modal,
        ):
            self._cancel_evt.set()
            if hasattr(self, "_step_lbl"):
                self._step_lbl.config(text="Cancelling after current step…")

    def _on_progress(self, payload):
        step, pct, detail = payload
        if not self._modal: return

        if step == "FlashingPhase":
            if self._modal_phase != "flashing":
                self._set_phase("flashing")
            remaining = max(0, 180 - pct)
            self._update_flashing(remaining)
            return

        if step == "FwIncompatible":
            self._set_phase("fw_compat_fail", detail=detail)
            return

        if step == "Reconnect":
            # Worker is waiting for the device to come back. pct = seconds
            # remaining (counts down to 0), detail = context-specific message.
            if self._modal_phase != "reconnecting":
                self._set_phase("reconnecting", detail=detail, secs_left=pct)
            else:
                self._update_reconnecting(pct, detail)
            return

        if pct < 0:
            if step == "Firmware":
                self._set_phase("fw_fail", detail=detail)
            elif hasattr(self, "_detail_lbl"):
                self._detail_lbl.config(text=detail, foreground="#c00")
            return

        self._update_step(step, pct, detail)

    def _on_device_done(self, result: DeviceResult):
        self._append_result(result)
        self.status_lbl.config(text=f"{len(self.results)} device(s) configured.")
        if not self._batch_active:
            self._close_modal()
            return
        # Reset per-device flags and morph the modal back to "waiting".
        self._cancel_evt.clear()
        self._skip_firmware_evt.clear()
        self._set_phase("waiting",
                        headline="Waiting for next device…",
                        sub="Unplug the current one and plug the next.")
        self._begin_detect(use_last_port=True)

    # ── End session / CSV / cross-thread drain ───────────────────────────
    def _end_batch(self):
        self._batch_active = False
        self._cancel_evt.set()
        self._close_modal()
        self._lock_form(False)
        self.status_lbl.config(text=f"Session ended. {len(self.results)} device(s).")

    def _reset_all(self):
        """Restore the app to its just-opened state: end any active batch,
        clear the grid, clear every form field. Confirms first if there
        are unexported results so the operator doesn't nuke unsaved data."""
        if self.results:
            if not messagebox.askyesno(
                "Reset?",
                f"This will clear the grid ({len(self.results)} row"
                f"{'s' if len(self.results) != 1 else ''}) and all form "
                f"fields.\n\nMake sure you've exported first.\n\nContinue?"):
                return
        # If a batch is running, end it cleanly first.
        if self._batch_active:
            self._end_batch()
        # Grid + results.
        self.results.clear()
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        # Form fields.
        self.do_fw_var.set(False);   self.fw_path_var.set("")
        self.do_cfg_var.set(False);  self.cfg_path_var.set("")
        self.group_name_var.set("")
        self.routing_ip_var.set("")
        self.routing_port_var.set("")
        self.cur_kw_var.set("")
        self.new_kw_var.set("")
        self.kw_action_var.set("none")
        # Push form-dependent widget states.
        self._on_form_change()
        # Status line.
        self.status_lbl.config(text="Ready.")

    def _save_xlsx(self):
        """XLSX export — Group / Routing IP / Routing Port are read from
        the form AT EXPORT TIME, not stamped at process time. This lets the
        operator fill those fields after the batch has run and still get
        them into the exported file, applied uniformly to every row.
        Timestamp, Port, and Remarks are deliberately omitted (they're
        operator diagnostics, kept in the grid but not useful downstream).
        """
        if not self.results:
            messagebox.showinfo("No data", "No devices configured yet.")
            return
        default = (f"teltonika_configured_devices_"
                   f"{datetime.now().strftime('%m%d%y%H%M%S')}.xlsx")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path: return

        # Pull the three "current form value" fields ONCE, at click time.
        group_now        = self.group_name_var.get().strip()
        routing_ip_now   = self.routing_ip_var.get().strip()
        routing_port_now = self.routing_port_var.get().strip()

        wb = Workbook()
        ws = wb.active
        ws.title = "Devices"

        headers = ("IMEI", "Model", "Firmware", "Config file",
                   "Keyword", "Name", "Group", "Routing IP", "Routing Port")
        ws.append(headers)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFF2DC", end_color="FFF2DC",
                                  fill_type="solid")
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"

        # Every row gets the same Group / Routing values (from the form now).
        # Name is left blank for the user to fill in later in Excel.
        for r in self.results:
            ws.append((r.imei, r.model, r.firmware, r.config_file,
                       r.keyword, "", group_now,
                       routing_ip_now, routing_port_now))

        widths = (18, 22, 22, 32, 12, 24, 18, 18, 12)
        from openpyxl.utils import get_column_letter
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("Save failed",
                                 f"Could not write the XLSX file:\n{e}")
            return
        messagebox.showinfo("Saved",
                            f"Saved {len(self.results)} row(s) to:\n{path}")

    def _drain_msgs(self):
        try:
            while True:
                kind, payload = self._msg_q.get_nowait()
                if   kind == "detect_sub":  self._update_waiting_sub(payload)
                elif kind == "detected":    self._on_device_detected(payload)
                elif kind == "progress":    self._on_progress(payload)
                elif kind == "device_done": self._on_device_done(payload)
                elif kind == "ask_keyword":
                    imei, box, evt = payload
                    self._set_phase("ask_keyword", imei=imei,
                                    result_box=box, evt=evt)
        except queue.Empty:
            pass
        self.root.after(50, self._drain_msgs)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    style = ttk.Style()
    try:
        style.theme_use("vista" if sys.platform == "win32" else "clam")
    except tk.TclError:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
